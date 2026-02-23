import openpyxl
import re
import json
from collections import Counter, defaultdict
from openpyxl.utils import range_boundaries, get_column_letter

def get_r1c1_formula(formula, cell):
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula
    def replace_ref(match):
        col_str = match.group(1)
        row_str = match.group(2)
        col_idx = 0
        for char in col_str:
            col_idx = col_idx * 26 + (ord(char.upper()) - ord('A') + 1)
        row_idx = int(row_str)
        row_off = row_idx - cell.row
        col_off = col_idx - cell.column
        return f"R[{row_off}]C[{col_off}]"
    r1c1 = re.sub(r'(?<![A-Z$])([A-Z]+)([0-9]+)(?![0-9])', replace_ref, formula)
    return r1c1

def group_cells(cells):
    if not cells: return ""
    if len(cells) == 1: return cells[0]
    parsed = []
    for c in cells:
        m = re.match(r"([A-Z]+)([0-9]+)", c)
        if m:
            col_str, row_str = m.groups()
            col_idx = 0
            for char in col_str:
                col_idx = col_idx * 26 + (ord(char.upper()) - ord('A') + 1)
            parsed.append((int(row_str), col_idx))
    if not parsed: return ", ".join(cells)
    parsed = sorted(list(set(parsed)))
    row_segments = defaultdict(list)
    for r, c in parsed:
        if not row_segments[r]:
            row_segments[r].append([c, c])
        else:
            last_seg = row_segments[r][-1]
            if c == last_seg[1] + 1:
                last_seg[1] = c
            else:
                row_segments[r].append([c, c])
    final_ranges = []
    sorted_rows = sorted(row_segments.keys())
    processed_segs = set()
    for i, r in enumerate(sorted_rows):
        for s_idx, seg in enumerate(row_segments[r]):
            if (r, s_idx) in processed_segs: continue
            start_row = r
            end_row = r
            for next_r_idx in range(i + 1, len(sorted_rows)):
                next_r = sorted_rows[next_r_idx]
                if next_r != end_row + 1: break
                found_match = False
                for ns_idx, nseg in enumerate(row_segments[next_r]):
                    if nseg == seg and (next_r, ns_idx) not in processed_segs:
                        end_row = next_r
                        processed_segs.add((next_r, ns_idx))
                        found_match = True
                        break
                if not found_match: break
            start_cell = f"{get_column_letter(seg[0])}{start_row}"
            if start_row == end_row and seg[0] == seg[1]:
                final_ranges.append(start_cell)
            else:
                end_cell = f"{get_column_letter(seg[1])}{end_row}"
                final_ranges.append(f"{start_cell}:{end_cell}")
            processed_segs.add((r, s_idx))
    return ", ".join(final_ranges)

def run_audit(file_path):
    wb_val = openpyxl.load_workbook(file_path, data_only=True)
    wb_form = openpyxl.load_workbook(file_path, data_only=False)
    findings = []
    error_codes = ["#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!"]
    for sheet_name in wb_val.sheetnames:
        ws_v = wb_val[sheet_name]
        ws_f = wb_form[sheet_name]
        for row_v, row_f in zip(ws_v.iter_rows(), ws_f.iter_rows()):
            for cell_v, cell_f in zip(row_v, row_f):
                if cell_v.value in error_codes:
                    findings.append({
                        "Sheet Name": sheet_name, "Cell Reference": cell_v.coordinate,
                        "Description of the Location": f"Cell in {sheet_name}",
                        "Short Error Categories": "Calculation Error",
                        "Long Description of error": f"Cell evaluates to {cell_v.value} error."
                    })
                if isinstance(cell_f.value, str) and "#REF!" in cell_f.value:
                    findings.append({
                        "Sheet Name": sheet_name, "Cell Reference": cell_f.coordinate,
                        "Description of the Location": f"Formula in {sheet_name}",
                        "Short Error Categories": "Broken Reference",
                        "Long Description of error": f"Formula contains #REF! error."
                    })
    for name, defn in wb_form.defined_names.items():
        val = str(defn.value) if hasattr(defn, 'value') else str(defn)
        if "#REF!" in val:
            findings.append({
                "Sheet Name": "Name Manager", "Cell Reference": "N/A",
                "Description of the Location": f"Named Range",
                "Short Error Categories": "Dead Name",
                "Long Description of error": f"Named range points to #REF! error."
            })
    calc_sheets = [s for s in wb_form.sheetnames if s.startswith('Calc_') or s in ['IS', 'CF', 'BS', 'KPIs']]
    for sheet_name in calc_sheets:
        ws = wb_form[sheet_name]
        for row in ws.iter_rows(min_row=4, max_col=121):
            row_cells = row[1:]
            r1c1_formulas = []
            valid_cells = []
            for cell in row_cells:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    r1c1_formulas.append(get_r1c1_formula(cell.value, cell))
                    valid_cells.append(cell)
            if len(r1c1_formulas) > 10:
                counts = Counter(r1c1_formulas)
                if len(counts) > 1:
                    most_common = counts.most_common(1)[0][0]
                    for cell, f_r1c1 in zip(valid_cells, r1c1_formulas):
                        if f_r1c1 != most_common:
                            if cell.column > 2:
                                findings.append({
                                    "Sheet Name": sheet_name, "Cell Reference": cell.coordinate,
                                    "Description of the Location": f"Calculation row in {sheet_name}",
                                    "Short Error Categories": "Formula Inconsistency",
                                    "Long Description of error": f"Formula pattern break (R1C1: {f_r1c1})."
                                })
            formulas_in_row = [c for c in row_cells if isinstance(c.value, str) and c.value.startswith('=')]
            nums_in_row = [c for c in row_cells if isinstance(c.value, (int, float)) and c.value != 0]
            if len(formulas_in_row) > 50 and len(nums_in_row) > 0:
                for cell in row_cells:
                    if isinstance(cell.value, (int, float)) and cell.value != 0:
                        findings.append({
                            "Sheet Name": sheet_name, "Cell Reference": cell.coordinate,
                            "Description of the Location": f"Formula row in {sheet_name}",
                            "Short Error Categories": "Hard-coded Value",
                            "Long Description of error": f"Hard-coded number found in a row that appears to be mostly formulas."
                        })
    for sheet_name in ['IS', 'CF', 'BS', 'KPIs', 'Annual_Summary']:
        ws = wb_val[sheet_name]
        for row in ws.iter_rows(max_col=1):
            cell = row[0]
            if isinstance(cell.value, str) and cell.value.strip():
                if not any(u in cell.value for u in ['$', 'HL', '%', 'AUD', 'Date', 'Period']):
                    has_data = False
                    for c in range(2, 5):
                        if isinstance(ws.cell(row=cell.row, column=c).value, (int, float)):
                            has_data = True; break
                    if has_data:
                        findings.append({
                            "Sheet Name": sheet_name, "Cell Reference": cell.coordinate,
                            "Description of the Location": f"Line items in {sheet_name}",
                            "Short Error Categories": "Missing Label",
                            "Long Description of error": f"Line items missing unit indicator ($, HL, %, etc.)."
                        })
    bs_ws = wb_val['BS']
    asset_row = 0; le_row = 0
    for r in range(1, bs_ws.max_row + 1):
        v = bs_ws.cell(row=r, column=1).value
        if v == 'TOTAL ASSETS': asset_row = r
        if v == 'TOTAL LIABILITIES & EQUITY': le_row = r
    if asset_row and le_row:
        for col in range(2, 122):
            a = bs_ws.cell(row=asset_row, column=col).value or 0
            le = bs_ws.cell(row=le_row, column=col).value or 0
            if abs(a - le) > 0.1:
                findings.append({
                    "Sheet Name": "BS", "Cell Reference": bs_ws.cell(row=asset_row, column=col).coordinate,
                    "Description of the Location": "Balance Sheet Balance",
                    "Short Error Categories": "Logical Flaw",
                    "Long Description of error": f"Balance sheet out of balance."
                })
    grouped = defaultdict(list)
    for f in findings:
        key = (f["Sheet Name"], f["Description of the Location"], f["Short Error Categories"], f["Long Description of error"])
        grouped[key].append(f["Cell Reference"])
    final_findings = []
    for key, cell_list in grouped.items():
        sheet, loc, cat, desc = key
        final_findings.append({
            "Sheet Name": sheet,
            "Cell Reference": group_cells(cell_list),
            "Description of the Location": loc,
            "Short Error Categories": cat,
            "Long Description of error": desc
        })
    return final_findings

if __name__ == "__main__":
    results = run_audit("Brewery_Financial_Model_10Y 1.xlsx")
    print("| Sheet Name | Cell Reference | Description of the Location | Short Error Categories | Long Description of error |")
    print("|------------|----------------|-----------------------------|------------------------|---------------------------|")
    for r in results:
        print(f"| {r['Sheet Name']} | {r['Cell Reference']} | {r['Description of the Location']} | {r['Short Error Categories']} | {r['Long Description of error']} |")
