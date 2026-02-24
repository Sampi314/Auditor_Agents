import openpyxl
import re
import json
from collections import Counter, defaultdict
from spellchecker import SpellChecker
from openpyxl.styles import PatternFill
from openpyxl.worksheet.formula import ArrayFormula

def get_color_str(color):
    if color is None: return "None"
    if color.type == 'rgb': return f"#{color.rgb[2:]}"
    elif color.type == 'theme': return f"Theme({color.theme})"
    elif color.type == 'indexed': return f"Indexed({color.indexed})"
    return "None"

def a1_to_r1c1(formula, cell_row, cell_col):
    if isinstance(formula, ArrayFormula): formula = formula.t
    formula = str(formula)
    def replace_ref(match):
        col_str, row_str = match.groups()
        col_abs = col_str.startswith('$')
        row_abs = row_str.startswith('$')
        col_name = col_str.replace('$', '')
        row_num = int(row_str.replace('$', ''))
        col_index = 0
        for char in col_name: col_index = col_index * 26 + (ord(char.upper()) - ord('A') + 1)
        if row_abs: r_part = f"R{row_num}"
        else:
            r_part = f"R[{row_num - cell_row}]"
            if r_part == "R[0]": r_part = "R"
        if col_abs: c_part = f"C{col_index}"
        else:
            c_part = f"C[{col_index - cell_col}]"
            if c_part == "C[0]": c_part = "C"
        return r_part + c_part
    pattern = r'(\$?[A-Z]+)(\$?[0-9]+)'
    return re.sub(pattern, replace_ref, formula)

def full_audit(filename):
    wb = openpyxl.load_workbook(filename, data_only=False)
    wb_data = openpyxl.load_workbook(filename, data_only=True)
    findings = []

    # 1. Sentry & Efficiency (Mega-Formula)
    error_values = ["#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#SPILL!", "#CALC!"]
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        ws_data = wb_data[sheetname]
        for row in ws.iter_rows():
            for cell in row:
                val = ws_data[cell.coordinate].value
                if str(val) in error_values:
                    findings.append({"Sheet Name": sheetname, "Cell Reference": cell.coordinate, "Description": "Calculation Error", "Category": "Calculation Error", "Long Description": f"🔴 HIGH: Cell contains error {val}"})
                if cell.data_type == 'f':
                    formula = str(cell.value)
                    if len(formula) > 4000:
                        # Efficiency Finding
                        findings.append({
                            "Sheet Name": sheetname,
                            "Cell Reference": cell.coordinate,
                            "Description": "Mega-Formula (Bad Practice)",
                            "Category": "Mega-Formula",
                            "Long Description": f"🔴 HIGH: Extremely long formula ({len(formula)} chars) detected. Flagged as CRITICAL BAD PRACTICE due to total lack of auditability. Formulas should be broken into modular steps."
                        })
                        # Architect Finding
                        findings.append({
                            "Sheet Name": sheetname,
                            "Cell Reference": cell.coordinate,
                            "Description": "Modularity Failure",
                            "Category": "Modularity Failure",
                            "Long Description": f"⚠️ MEDIUM: Calculation is condensed into a 'black box' Mega-Formula. Structural integrity is compromised as modifications risk breaking complex nested logic."
                        })

    # 2. Lingo
    spell = SpellChecker()
    spell.word_frequency.load_words(['EBITDA', 'COGS', 'CFADS', 'IRR', 'CAPEX', 'OPEX', 'GST', 'Corality', 'Excel', 'tutorial', 'modelling', 'Offsheet', 'OneLine', 'DSCR', 'Heng'])
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        for row in ws.iter_rows(max_row=30):
            for cell in row:
                if isinstance(cell.value, str) and not cell.value.startswith('=') and cell.column <= 5:
                    words = re.findall(r'\b[A-Za-z]{4,}\b', cell.value)
                    misspelled = spell.unknown(words)
                    for word in misspelled:
                        if word.lower() not in ['dscr', 'heng']:
                            findings.append({"Sheet Name": sheetname, "Cell Reference": cell.coordinate, "Description": f"Label: {cell.value}", "Category": "Typo", "Long Description": f"🟡 LOW: Potential typo '{word}' in '{cell.value}'"})

    # 3. Stylist
    stylist_findings = defaultdict(list)
    for sheetname in wb.sheetnames:
        if sheetname in ['L', '  ', 'Corality']: continue
        ws = wb[sheetname]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None: continue
                f_color = get_color_str(cell.font.color)
                if not cell.data_type == 'f' and isinstance(cell.value, (int, float)):
                    if cell.row > 5 and cell.column > 3:
                        if f_color != "#800000" and f_color != "None" and f_color != "Theme(0)":
                            stylist_findings[(sheetname, "Colour Coding Error", f"🟡 LOW: Hard-coded input is formatted with font color {f_color} instead of the model's Input style (#800000)")].append(cell.coordinate)
                if cell.data_type == 'f' and '!' in str(cell.value):
                    if f_color != "Indexed(16)":
                        stylist_findings[(sheetname, "Colour Coding Error", "🟡 LOW: Off-sheet link is formatted with font color " + f_color + " instead of the model's Link style (Indexed 16)")].append(cell.coordinate)
    for (sn, cat, desc), cells in stylist_findings.items():
        findings.append({"Sheet Name": sn, "Cell Reference": ", ".join(cells), "Description": "Multiple locations", "Category": cat, "Long Description": desc})

    # 4. Logic & Deep Correctness
    for sheetname in wb.sheetnames:
        if sheetname in ['L', '  ', 'Corality']: continue
        ws = wb[sheetname]
        for row in ws.iter_rows():
            formulas = {}
            for cell in row:
                if cell.data_type == 'f':
                    try:
                        r1c1 = a1_to_r1c1(cell.value, cell.row, cell.column)
                        formulas[cell.coordinate] = r1c1

                        # Deep Correctness for Mega-Formulas
                        if len(str(cell.value)) > 4000:
                            # If it doesn't result in an error, and follow the dominant row pattern, it's "technically correct" but still bad.
                            val = wb_data[sheetname][cell.coordinate].value
                            if val not in error_values:
                                findings.append({
                                    "Sheet Name": sheetname,
                                    "Cell Reference": cell.coordinate,
                                    "Description": "Deep Logic Check",
                                    "Category": "Auditability Risk",
                                    "Long Description": f"⚠️ MEDIUM: Mega-formula is technically functioning (results in {val}), but its complexity makes it an extreme auditability risk."
                                })
                    except: pass

            if len(formulas) > 5:
                counts = Counter(formulas.values())
                if len(counts) > 1:
                    dominant_r1c1, _ = counts.most_common(1)[0]
                    for coord, r1c1 in formulas.items():
                        if r1c1 != dominant_r1c1 and ws[coord].column >= 9:
                            findings.append({"Sheet Name": sheetname, "Cell Reference": coord, "Description": "Formula row", "Category": "Formula Pattern Break", "Long Description": f"⚠️ MEDIUM: Formula differs from dominant pattern in row. Expected R1C1: {dominant_r1c1} vs Actual R1C1: {r1c1}"})

        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    formula = str(cell.value)
                    literals = re.findall(r'(?<![A-Z$])\b([0-9]+\.[0-9]+|[2-9][0-9]+|[0-9]{2,})\b', formula)
                    literals = [l for l in literals if l not in ['100', '12', '365', '52', '1', '0', '2', '4']]
                    if literals:
                        findings.append({"Sheet Name": sheetname, "Cell Reference": cell.coordinate, "Description": "Hard-coded literal", "Category": "Hard-Code in Formula", "Long Description": f"⚠️ MEDIUM: Formula contains hard-coded literal(s) {literals} which should likely be cell references."})

    return findings

if __name__ == "__main__":
    results = full_audit('20130401 Efficient Modelling (Tutorial).xlsx')
    with open('full_audit_results.json', 'w') as f: json.dump(results, f)
