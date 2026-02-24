import openpyxl
import re
import os
import sys
import argparse
from collections import defaultdict, Counter

def get_sheet_role(sheetname):
    s = sheetname.lower()
    if any(x in s for x in ['cover', 'toc', 'index', 'navigation']): return "Cover / TOC"
    if any(x in s for x in ['control', 'scenario', 'switch']): return "Control / Scenario"
    if any(x in s for x in ['assumption', 'input', 'driver']): return "Inputs / Assumptions"
    if any(x in s for x in ['timing', 'timeline', 'date']): return "Timing"
    if any(x in s for x in ['calc', 'revenue', 'opex', 'capex', 'debt', 'tax', 'deprn']): return "Calculations"
    if any(x in s for x in ['p&l', 'bs', 'cfs', 'is', 'statement', 'cashflow']): return "Financial Statements"
    if any(x in s for x in ['summary', 'dashboard', 'output', 'kpi']): return "Outputs / Dashboard"
    if any(x in s for x in ['check', 'balance', 'audit']): return "Checks"
    if any(x in s for x in ['data', 'lookup', 'table']): return "Data / Lookup"
    return "Other"

def parse_formula_references(formula, current_sheet, sheet_names):
    """Extracts (sheet, cell) references from an Excel formula."""
    refs = set()
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return refs

    # 1. Handle cross-sheet references: 'Sheet Name'!A1 or SheetName!A1
    # Handle ranges too: Sheet1!A1:B10
    cross_sheet_pattern = r"('([^']+)'|([a-zA-Z0-9_.& ]+))!(\$?[A-Z]+\$?[0-9]+)(:(\$?[A-Z]+\$?[0-9]+))?"
    for match in re.finditer(cross_sheet_pattern, formula):
        quoted_name = match.group(2)
        unquoted_name = match.group(3)
        sheet_name = quoted_name if quoted_name else unquoted_name
        cell_start = match.group(4).replace('$', '')
        cell_end = match.group(6).replace('$', '') if match.group(6) else None

        if sheet_name in sheet_names:
            refs.add((sheet_name, cell_start))
            if cell_end:
                refs.add((sheet_name, cell_end))

    # 2. Handle same-sheet references
    # First, strip cross-sheet references to avoid double counting
    stripped_formula = re.sub(cross_sheet_pattern, " ", formula)
    # Also strip strings and functions
    stripped_formula = re.sub(r'"[^"]*"', " ", stripped_formula)

    # Same sheet cell pattern
    same_sheet_pattern = r"(?<![A-Z0-9_.])(\$?[A-Z]+\$?[0-9]+)(?![A-Z0-9_.])"
    for match in re.finditer(same_sheet_pattern, stripped_formula):
        cell_ref = match.group(1).replace('$', '')
        # Basic validation: ensure it's not a number
        if not cell_ref.isdigit():
            refs.add((current_sheet, cell_ref))

    return refs

def get_row_labels(ws):
    """Attempts to find labels for each row in the first few columns."""
    labels = {}
    for row in ws.iter_rows(min_col=1, max_col=5, max_row=min(ws.max_row, 500)):
        label = ""
        for cell in row:
            if isinstance(cell.value, str) and not str(cell.value).startswith('='):
                val = str(cell.value).strip()
                if val:
                    label = val
                    labels[cell.row] = label
                    break
    return labels

def detect_sections(ws, labels):
    """Detects sections based on row formatting and label patterns."""
    sections = []

    # A simple heuristic: rows that have labels in column 1 but no formulas in first 10 columns
    # might be section headers.

    labeled_rows = sorted(labels.keys())
    for i, row_idx in enumerate(labeled_rows):
        label = labels[row_idx]

        # Heuristic for header: Uppercase, or bold, or no formula in row
        is_header = False
        if label.isupper() and len(label) > 3:
            is_header = True

        if not is_header:
            # Check for bold
            cell = ws.cell(row=row_idx, column=1)
            if cell.font and cell.font.bold:
                is_header = True

        if is_header:
            # Check if it has any formula in columns 2-10
            has_formula = False
            for col_idx in range(2, 11):
                c = ws.cell(row=row_idx, column=col_idx)
                if c.data_type == 'f':
                    has_formula = True
                    break
            if not has_formula:
                sections.append((row_idx, label))

    if not sections:
        sections = [(1, "Main")]
    return sections

def generate_cartographer_reports(filename, output_dir=None):
    print(f"Loading {filename}...")
    wb = openpyxl.load_workbook(filename, data_only=False)
    # Also load data_only for value checks if needed, but we mostly need formulas
    sheet_names = wb.sheetnames

    graph = defaultdict(set)
    sheet_info = {}

    for sheetname in sheet_names:
        print(f"Scanning sheet: {sheetname}")
        ws = wb[sheetname]
        role = get_sheet_role(sheetname)
        labels = get_row_labels(ws)
        sections = detect_sections(ws, labels)

        sheet_info[sheetname] = {
            'role': role,
            'visible': ws.sheet_state == 'visible',
            'sections': sections,
            'labels': labels
        }

        # Scan formulas
        for row in ws.iter_rows(max_row=300, max_col=15): # Scan subset for performance/relevance
            for cell in row:
                if cell.data_type == 'f':
                    formula = str(cell.value)
                    refs = parse_formula_references(formula, sheetname, sheet_names)
                    target = (sheetname, cell.coordinate)
                    for ref in refs:
                        graph[target].add(ref)

    # 1. Flow_Dependency_Register.md
    print("Generating Dependency Register...")
    register_lines = ["# Flow Dependency Register", "", "| # | Source Sheet | Source Cell | Source Row Label | Target Sheet | Target Cell | Target Row Label | Reference Type |", "|---|---|---|---|---|---|---|---|"]

    count = 1
    processed_links = set()

    # Sort targets to make report deterministic
    for target in sorted(graph.keys()):
        sources = graph[target]
        t_sheet, t_cell = target
        t_row_match = re.search(r'\d+', t_cell)
        if not t_row_match: continue
        t_row = int(t_row_match.group())
        t_label = sheet_info[t_sheet]['labels'].get(t_row, "")

        for s_sheet, s_cell in sorted(sources):
            if s_sheet != t_sheet:
                s_row_match = re.search(r'\d+', s_cell)
                if not s_row_match: continue
                s_row = int(s_row_match.group())
                s_label = sheet_info[s_sheet]['labels'].get(s_row, "")

                link_key = (s_sheet, s_label, t_sheet, t_label)
                if link_key in processed_links: continue
                processed_links.add(link_key)

                ref_type = "LINK"
                s_ws = wb[s_sheet]
                if s_ws[s_cell].data_type != 'f':
                    ref_type = "INPUT"

                register_lines.append(f"| {count} | {s_sheet} | {s_cell} | {s_label} | {t_sheet} | {t_cell} | {t_label} | {ref_type} |")
                count += 1

    # Ensure Maps directory exists
    if output_dir:
        maps_dir = os.path.join(output_dir, 'Maps')
    else:
        maps_dir = 'Maps'

    if not os.path.exists(maps_dir):
        os.makedirs(maps_dir)

    if output_dir:
        reg_path = os.path.join(maps_dir, 'Flow_Dependency_Register.md')
    else:
        reg_path = os.path.join(maps_dir, 'Flow_Dependency_Register.md')

    with open(reg_path, 'w') as f:
        f.write("\n".join(register_lines))

    # 2. Flow_L1_Workbook.mermaid
    print("Generating L1 Workbook Map...")
    l1_lines = ["flowchart LR", ""]

    roles = defaultdict(list)
    for s in sheet_names:
        roles[sheet_info[s]['role']].append(s)

    for role, sheets in roles.items():
        role_id = "sg_" + role.replace(' ', '_').replace('/', '_').replace('&', 'and')
        l1_lines.append(f"    subgraph {role_id}[\"{role}\"]")
        for s in sheets:
            s_id = s.replace(' ', '_').replace('&', 'and').replace('-', '_').replace('!', '_')
            label = s
            if not sheet_info[s]['visible']: label += " (hidden)"
            l1_lines.append(f"        {s_id}[\"{label}\"]")
        l1_lines.append("    end")

    sheet_conns = defaultdict(lambda: defaultdict(set))
    for target, sources in graph.items():
        t_sheet, t_cell = target
        for s_sheet, s_cell in sources:
            if s_sheet != t_sheet:
                s_row_match = re.search(r'\d+', s_cell)
                if s_row_match:
                    s_row = int(s_row_match.group())
                    s_label = sheet_info[s_sheet]['labels'].get(s_row, "")
                    if s_label:
                        sheet_conns[s_sheet][t_sheet].add(s_label)

    for src, dsts in sorted(sheet_conns.items()):
        src_id = src.replace(' ', '_').replace('&', 'and').replace('-', '_').replace('!', '_')
        for dst, labels in sorted(dsts.items()):
            dst_id = dst.replace(' ', '_').replace('&', 'and').replace('-', '_').replace('!', '_')
            label_list = sorted(list(labels))
            if len(label_list) > 3:
                label_str = ", ".join(label_list[:3]) + f" (+{len(label_list)-3} more)"
            else:
                label_str = ", ".join(label_list)
            l1_lines.append(f"    {src_id} -- \"{label_str}\" --> {dst_id}")

    with open(os.path.join(maps_dir, 'Flow_L1_Workbook.mermaid'), 'w') as f:
        f.write("\n".join(l1_lines))

    # 3. Flow_L2_{SheetName}.mermaid
    for sheetname in sheet_names:
        print(f"Generating L2 Map for {sheetname}...")
        l2_lines = ["flowchart TD", ""]

        sections = sheet_info[sheetname]['sections']
        s_id_base = sheetname.replace(' ', '_').replace('&', 'and').replace('-', '_').replace('!', '_')

        l2_lines.append(f"    subgraph sg_{s_id_base}[\"{sheetname}\"]")
        for i, (s_row, s_label) in enumerate(sections):
            sect_id = f"sect_{s_id_base}_{i}"
            l2_lines.append(f"        {sect_id}[\"{s_label}\"]")
        l2_lines.append("    end")

        ext_in = defaultdict(set)
        ext_out = defaultdict(set)
        int_conns = defaultdict(set)

        for target, sources in graph.items():
            t_sheet, t_cell = target
            t_row_match = re.search(r'\d+', t_cell)
            if not t_row_match: continue
            t_row = int(t_row_match.group())

            t_sect_idx = 0
            for i, (sec_row, sec_label) in enumerate(sections):
                if t_row >= sec_row: t_sect_idx = i

            if t_sheet == sheetname:
                for s_sheet, s_cell in sources:
                    s_row_match = re.search(r'\d+', s_cell)
                    if not s_row_match: continue
                    s_row = int(s_row_match.group())
                    s_label = sheet_info[s_sheet]['labels'].get(s_row, "Data")

                    if s_sheet != sheetname:
                        ext_in[s_sheet].add((s_label, t_sect_idx))
                    else:
                        s_sect_idx = 0
                        for i, (sec_row, sec_label) in enumerate(sections):
                            if s_row >= sec_row: s_sect_idx = i
                        if s_sect_idx != t_sect_idx:
                            int_conns[(s_sect_idx, t_sect_idx)].add(s_label)
            else:
                for s_sheet, s_cell in sources:
                    if s_sheet == sheetname:
                        s_row_match = re.search(r'\d+', s_cell)
                        if not s_row_match: continue
                        s_row = int(s_row_match.group())
                        s_sect_idx = 0
                        for i, (sec_row, sec_label) in enumerate(sections):
                            if s_row >= sec_row: s_sect_idx = i

                        s_label = sheet_info[s_sheet]['labels'].get(s_row, "Data")
                        ext_out[t_sheet].add((s_label, s_sect_idx))

        for ext_sheet, connections in sorted(ext_in.items()):
            ext_id = "ext_in_" + ext_sheet.replace(' ', '_').replace('&', 'and').replace('-', '_').replace('!', '_')
            l2_lines.append(f"    {ext_id}([\"{ext_sheet}\"])")
            by_sect = defaultdict(list)
            for lab, s_idx in sorted(connections): by_sect[s_idx].append(lab)
            for s_idx, labels in sorted(by_sect.items()):
                sect_id = f"sect_{s_id_base}_{s_idx}"
                label_str = ", ".join(list(dict.fromkeys(labels))[:2])
                if len(set(labels)) > 2: label_str += "..."
                l2_lines.append(f"    {ext_id} -- \"{label_str}\" --> {sect_id}")

        for (src_idx, dst_idx), labels in sorted(int_conns.items()):
            src_sect_id = f"sect_{s_id_base}_{src_idx}"
            dst_sect_id = f"sect_{s_id_base}_{dst_idx}"
            label_str = ", ".join(list(dict.fromkeys(labels))[:2])
            if len(set(labels)) > 2: label_str += "..."
            l2_lines.append(f"    {src_sect_id} -- \"{label_str}\" --> {dst_sect_id}")

        for ext_sheet, connections in sorted(ext_out.items()):
            ext_id = "ext_out_" + ext_sheet.replace(' ', '_').replace('&', 'and').replace('-', '_').replace('!', '_')
            l2_lines.append(f"    {ext_id}([\"{ext_sheet}\"])")
            by_sect = defaultdict(list)
            for lab, s_idx in sorted(connections): by_sect[s_idx].append(lab)
            for s_idx, labels in sorted(by_sect.items()):
                sect_id = f"sect_{s_id_base}_{s_idx}"
                label_str = ", ".join(list(dict.fromkeys(labels))[:2])
                if len(set(labels)) > 2: label_str += "..."
                l2_lines.append(f"    {sect_id} -- \"{label_str}\" --> {ext_id}")

        filename_l2 = os.path.join(maps_dir, f'Flow_L2_{sheetname.replace(" ", "_")}.mermaid')
        with open(filename_l2, 'w') as f:
            f.write("\n".join(l2_lines))

    # 4. Flow_L3_Critical_Path.mermaid
    print("Generating L3 Critical Path Map...")
    l3_lines = ["flowchart LR", "", "    classDef input fill:#E8F5E9,stroke:#2E7D32", "    classDef shared fill:#FFD700,stroke:#333", "    classDef output fill:#E3F2FD,stroke:#1565C0", ""]

    key_outputs = []
    for sn in ['KPIs', 'Annual_Summary', 'Summary']:
        if sn in sheet_names:
            labels = sheet_info[sn]['labels']
            for row, label in labels.items():
                if any(x in label.upper() for x in ['NET INCOME', 'EBITDA', 'IRR', 'DSCR', 'NPV', 'CASH BALANCE']):
                    # Check col D, E, or F for first period
                    for col in [4, 5, 6]:
                        cell = wb[sn].cell(row=row, column=col)
                        if cell.data_type == 'f':
                            key_outputs.append((sn, cell.coordinate, label))
                            break

    nodes = set()
    edges = set()
    visited = set()

    def trace(target, depth):
        if depth > 3 or target in visited: return
        visited.add(target)
        t_sheet, t_cell = target
        t_row_match = re.search(r'\d+', t_cell)
        if not t_row_match: return
        t_row = int(t_row_match.group())
        t_label = sheet_info[t_sheet]['labels'].get(t_row, f"{t_sheet}!{t_cell}")

        t_id = (t_sheet + "_" + t_cell).replace(' ', '_').replace('&', 'and').replace('-', '_').replace('!', '_')
        is_output = depth == 0

        sources = graph.get(target, [])
        is_input = len(sources) == 0 or wb[t_sheet][t_cell].data_type != 'f'

        nodes.add((t_id, t_label, is_output, is_input))

        for s_sheet, s_cell in sources:
            s_row_match = re.search(r'\d+', s_cell)
            if not s_row_match: continue
            s_row = int(s_row_match.group())
            s_label = sheet_info[s_sheet]['labels'].get(s_row, f"{s_sheet}!{s_cell}")
            s_id = (s_sheet + "_" + s_cell).replace(' ', '_').replace('&', 'and').replace('-', '_').replace('!', '_')

            s_sources = graph.get((s_sheet, s_cell), [])
            s_is_input = len(s_sources) == 0 or wb[s_sheet][s_cell].data_type != 'f'

            nodes.add((s_id, s_label, False, s_is_input))
            edges.add((s_id, t_id))

            if not s_is_input:
                trace((s_sheet, s_cell), depth + 1)

    for sn, coord, label in key_outputs[:5]:
        trace((sn, coord), 0)

    for n in sorted(list(nodes)):
        n_id, n_label, is_out, is_in = n
        n_label = n_label.replace('"', "'")
        style = ""
        if is_out: style = ":::output"
        elif is_in: style = ":::input"
        l3_lines.append(f"    {n_id}[\"{n_label}\"]{style}")

    for s_id, t_id in sorted(list(edges)):
        l3_lines.append(f"    {s_id} --> {t_id}")

    with open(os.path.join(maps_dir, 'Flow_L3_Critical_Path.mermaid'), 'w') as f:
        f.write("\n".join(l3_lines))

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Cartographer Script")
    parser.add_argument("--input", required=True, help="Path to input Excel file")
    parser.add_argument("--output-dir", required=True, help="Path to output directory")
    args = parser.parse_args()

    if not os.path.exists(args.output_dir):
        os.makedirs(args.output_dir)

    generate_cartographer_reports(args.input, args.output_dir)
