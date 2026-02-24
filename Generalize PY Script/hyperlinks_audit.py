import openpyxl
import json
import re
import os
import argparse

def add_finding(findings, sheetname, ref, desc, cat, long_desc, priority, agent):
    findings.append({
        "Sheet Name": sheetname,
        "Cell Reference": ref,
        "Description": desc,
        "Category": cat,
        "Long Description": long_desc,
        "Priority": priority,
        "Agent": agent
    })

def audit_hyperlinks(filename, output_dir=None):
    wb = openpyxl.load_workbook(filename, data_only=True)
    findings = []

    # Load with formulas
    wb_form = openpyxl.load_workbook(filename, data_only=False)

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        ws_form = wb_form[sheetname]

        # Check formulas for HYPERLINK
        for row in ws_form.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and 'HYPERLINK' in str(cell.value).upper():
                    formula = str(cell.value)
                    # Simple regex to extract link from HYPERLINK("link", "friendly")
                    match = re.search(r'HYPERLINK\("([^"]+)"', formula, re.IGNORECASE)
                    if match:
                        link_target = match.group(1)
                        if link_target.startswith('#'):
                            # Internal link in formula
                            location = link_target[1:]
                            ref = cell.coordinate
                            if '!' in location:
                                t_sn, t_cr = location.split('!')
                                t_sn = t_sn.strip("'")
                                if t_sn not in wb.sheetnames:
                                    add_finding(findings, sheetname, ref, f"HYPERLINK formula to {link_target}", "Broken Link - Missing Sheet",
                                                f"HYPERLINK formula references sheet '{t_sn}' which does not exist. Update the formula to reference a valid sheet.", "High", "Hyperlinks")
                        else:
                            # External link in formula
                            ref = cell.coordinate
                            if "example.com" in link_target:
                                add_finding(findings, sheetname, ref, f"HYPERLINK formula to {link_target}", "Placeholder URL",
                                            "External URL in formula contains obvious placeholder text. Update with the correct destination URL.", "Low", "Hyperlinks")

        # openpyxl stores hyperlinks in ws._hyperlinks
        for rel in ws._hyperlinks:
            target = rel.target
            location = rel.location # e.g. 'Sheet1!A1'
            ref = rel.ref # e.g. 'A1'

            # Internal link
            if location:
                # Check if it's a sheet reference
                if '!' in location:
                    target_sheetname, target_cell_ref = location.split('!')
                    target_sheetname = target_sheetname.strip("'")
                    if target_sheetname not in wb.sheetnames:
                        add_finding(findings, sheetname, ref, f"Hyperlink to {location}", "Broken Link - Missing Sheet",
                                    f"Hyperlink references sheet '{target_sheetname}' which does not exist. Check for renamed or deleted sheets.", "High", "Hyperlinks")
                    else:
                        # Check if target cell is blank
                        target_ws = wb[target_sheetname]
                        try:
                            target_val = target_ws[target_cell_ref].value
                            if target_val is None or str(target_val).strip() == "":
                                add_finding(findings, sheetname, ref, f"Hyperlink to {location}", "Link to Blank Cell",
                                            "Hyperlink navigates to a valid cell, but the destination cell is blank. Ensure the target contains relevant content or update the link.", "Medium", "Hyperlinks")
                        except:
                            # Might be out of bounds or invalid ref
                            add_finding(findings, sheetname, ref, f"Hyperlink to {location}", "Broken Link - Out of Bounds",
                                        "Hyperlink targets a cell reference beyond the sheet's bounds or is invalid. Correct the target reference.", "High", "Hyperlinks")
                else:
                    # Likely a named range or local cell
                    if location not in wb.defined_names:
                        # Try if it's a cell ref in the same sheet
                        try:
                            target_val = ws[location].value
                            if target_val is None or str(target_val).strip() == "":
                                add_finding(findings, sheetname, ref, f"Hyperlink to {location}", "Link to Blank Cell",
                                            "Hyperlink navigates to a valid cell, but the destination cell is blank. Ensure the target contains relevant content or update the link.", "Medium", "Hyperlinks")
                        except:
                             add_finding(findings, sheetname, ref, f"Hyperlink to {location}", "Broken Link - Missing Name",
                                         f"Hyperlink targets a named range or cell '{location}' which does not exist. Verify the target name or reference.", "High", "Hyperlinks")

            # External link
            if target and not location:
                if "example.com" in target or "xxx" in target:
                    add_finding(findings, sheetname, ref, f"Hyperlink to {target}", "Placeholder URL",
                                "External URL contains obvious placeholder text. Update with the correct destination URL.", "Low", "Hyperlinks")
                if target.startswith("C:\\Users\\"):
                    add_finding(findings, sheetname, ref, f"Hyperlink to {target}", "Non-Portable File Path",
                                "File path is user-specific and will break for other users. Use relative paths or a shared network location.", "Medium", "Hyperlinks")

    # Write findings
    if output_dir:
        output_path = os.path.join(output_dir, 'full_audit_results.json')
    else:
        output_path = 'full_audit_results.json'

    # Read existing
    all_findings = []
    if os.path.exists(output_path):
        try:
            with open(output_path, 'r') as f:
                all_findings = json.load(f)
        except:
            pass

    all_findings.extend(findings)

    with open(output_path, 'w') as f:
        json.dump(all_findings, f)

    return findings

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Hyperlinks Audit Script")
    parser.add_argument("--input", required=True, help="Path to input Excel file")
    parser.add_argument("--output-dir", required=True, help="Path to output directory")
    args = parser.parse_args()

    if not os.path.exists(args.output_dir):
        os.makedirs(args.output_dir)

    print(f"Auditing Hyperlinks: {args.input}")
    audit_hyperlinks(args.input, args.output_dir)
