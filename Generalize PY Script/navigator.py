import openpyxl
import re
import os
import argparse
from collections import defaultdict

def get_sheet_role(sheetname):
    s = sheetname.lower()
    if any(x in s for x in ['cover', 'toc', 'index', 'navigation']): return "Cover / TOC"
    if any(x in s for x in ['control', 'scenario', 'switch']): return "Control / Scenario"
    if any(x in s for x in ['assumption', 'input', 'driver']): return "Inputs / Assumptions"
    if any(x in s for x in ['timing', 'timeline', 'date']): return "Timing"
    if any(x in s for x in ['calc', 'revenue', 'opex', 'capex', 'debt', 'tax', 'deprn']): return "Calculations"
    if any(x in s for x in ['p&l', 'bs', 'cfs', 'is', 'statement', 'cashflow']): return "Financial Statements"
    if any(x in s for x in ['summary', 'dashboard', 'output', 'kpi']): return "Outputs / Dashboard"
    if any(x in s for x in ['check', 'balance', 'audit']): return "Checks"
    if any(x in s for x in ['data', 'lookup', 'table']): return "Data / Lookup"
    return "Other"

def generate_navigator_reports(filename, output_dir=None):
    wb = openpyxl.load_workbook(filename, data_only=False)
    wb_data = openpyxl.load_workbook(filename, data_only=True)

    # 1. Model Summary
    summary_lines = ["# 01 Model Summary", ""]
    summary_lines.append("## 1. Purpose")
    summary_lines.append("Automated summary of the financial model.")
    summary_lines.append("")
    summary_lines.append("## 2. Model Structure")
    summary_lines.append("| Sheet Name | Role | Description |")
    summary_lines.append("|---|---|---|")

    for sheetname in wb.sheetnames:
        role = get_sheet_role(sheetname)
        summary_lines.append(f"| {sheetname} | {role} | Financial model component |")

    summary_lines.append("")

    # Timeline
    summary_lines.append("## 3. Timeline")
    timeline_sheet = next((s for s in wb.sheetnames if any(x in s.lower() for x in ['time', 'date', 'timing'])), None)
    if timeline_sheet:
        summary_lines.append(f"Timeline detected in sheet: {timeline_sheet}")
    else:
        summary_lines.append("No explicit timeline sheet detected.")
    summary_lines.append("")

    if output_dir:
        summary_path = os.path.join(output_dir, '01_Model_Summary.md')
    else:
        summary_path = '01_Model_Summary.md'

    with open(summary_path, 'w') as f:
        f.write("\n".join(summary_lines))

    # 2. Model Flowchart (Mermaid)
    flow_lines = ["# 02 Model Flowchart", "", "```mermaid", "flowchart LR"]

    # Group sheets by role for subgraphs
    roles = defaultdict(list)
    for s in wb.sheetnames:
        roles[get_sheet_role(s)].append(s)

    for role, sheets in roles.items():
        flow_lines.append(f"    subgraph {role.replace(' ', '_')}")
        for s in sheets:
            clean_s = s.replace(' ', '_').replace('&', 'and')
            flow_lines.append(f"        {clean_s}[{s}]")
        flow_lines.append("    end")

    # Detect connections
    connections = set()
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        for row in ws.iter_rows(max_row=100, max_col=20): # Scan subset for performance
            for cell in row:
                if cell.data_type == 'f':
                    formula = str(cell.value)
                    # Find sheet references like 'Sheet Name'! or SheetName!
                    refs = re.findall(r"'?([a-zA-Z0-9 _&]+)'?!", formula)
                    for r in refs:
                        if r in wb.sheetnames and r != sheetname:
                            connections.add((r, sheetname))

    for src, dst in connections:
        src_c = src.replace(' ', '_').replace('&', 'and')
        dst_c = dst.replace(' ', '_').replace('&', 'and')
        flow_lines.append(f"    {src_c} --> {dst_c}")

    flow_lines.append("```")

    if output_dir:
        flow_path = os.path.join(output_dir, '02_Model_Flowchart.md')
    else:
        flow_path = '02_Model_Flowchart.md'

    with open(flow_path, 'w') as f:
        f.write("\n".join(flow_lines))

    # 3. Calculation Reference
    calc_lines = ["# 03 Calculation Reference", ""]
    calc_lines.append("## Key Calculations")
    calc_lines.append("")
    calc_lines.append("| Sheet | Row | Plain English | Readable Formula (RFL) | A1 Reference |")
    calc_lines.append("|---|---|---|---|---|")

    # Just take some representative calculations from Calc sheets
    calc_sheets = [s for s in wb.sheetnames if get_sheet_role(s) == "Calculations"]
    for s in calc_sheets[:3]: # Limit to first 3 calc sheets
        ws = wb[s]
        ws_data = wb_data[s]
        count = 0
        for row in ws.iter_rows(min_row=5, max_row=100):
            # Find the first non-empty cell in the first few columns to use as label
            label = ""
            for i in range(5):
                if row[i].value and row[i].data_type != 'f':
                    label = str(row[i].value).strip()
                    break
            if not label or label == "": continue

            for cell in row[1:15]: # Check first few cols
                if cell.data_type == 'f':
                    formula = str(cell.value)
                    calc_lines.append(f"| {s} | {cell.row} | {label} | `{formula}` | `{cell.coordinate}` |")
                    count += 1
                    break
            if count >= 5: break # Max 5 per sheet

    if output_dir:
        calc_path = os.path.join(output_dir, '03_Calculation_Reference.md')
    else:
        calc_path = '03_Calculation_Reference.md'

    with open(calc_path, 'w') as f:
        f.write("\n".join(calc_lines))

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Navigator Script")
    parser.add_argument("--input", required=True, help="Path to input Excel file")
    parser.add_argument("--output-dir", required=True, help="Path to output directory")
    args = parser.parse_args()

    if not os.path.exists(args.output_dir):
        os.makedirs(args.output_dir)

    print(f"Generating Navigator reports for: {args.input}")
    generate_navigator_reports(args.input, args.output_dir)
