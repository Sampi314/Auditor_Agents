import openpyxl

def mega_formula_audit(filename, threshold=4000):
    wb = openpyxl.load_workbook(filename, data_only=False)
    findings = []

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    formula = str(cell.value)
                    if len(formula) > threshold:
                        findings.append({
                            "Sheet Name": sheetname,
                            "Cell Reference": cell.coordinate,
                            "Length": len(formula)
                        })

    # Group by sheet and similar length/formula pattern if possible
    # For now, just list all as requested.
    return findings

if __name__ == "__main__":
    findings = mega_formula_audit('20130401 Efficient Modelling (Tutorial).xlsx')
    import json
    final_findings = []

    from collections import defaultdict
    grouped = defaultdict(list)
    for f in findings:
        grouped[(f["Sheet Name"], f["Length"])].append(f["Cell Reference"])

    for (sn, length), cells in grouped.items():
        final_findings.append({
            "Sheet Name": sn,
            "Cell Reference": ", ".join(cells),
            "Description": f"Formula with {length} characters",
            "Category": "Mega-Formula",
            "Long Description": f"🔴 HIGH: Extremely long formula ({length} chars) detected. Mega-formulas are impossible to audit or maintain."
        })

    with open('findings_mega.json', 'w') as f:
        json.dump(final_findings, f)
