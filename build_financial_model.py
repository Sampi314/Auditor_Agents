#!/usr/bin/env python3
"""
3-Statement Financial Model Builder with Control Accounts
Builds a fully integrated financial model in Excel with proper financial modeling standards
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import calendar

def create_financial_model():
    """Create a comprehensive 3-statement financial model with control accounts"""

    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Define constants
    NUM_MONTHS = 12
    START_MONTH = 1  # January
    START_YEAR = 2025

    # Create sheets in logical order
    ws_inputs = wb.create_sheet("Inputs")
    ws_workings = wb.create_sheet("Workings")
    ws_income = wb.create_sheet("Income Statement")
    ws_balance = wb.create_sheet("Balance Sheet")
    ws_cashflow = wb.create_sheet("Cash Flow Statement")

    # Define color scheme for financial modeling best practice
    COLOR_HEADER = "366092"  # Dark blue
    COLOR_INPUT = "FFFF00"   # Yellow for inputs
    COLOR_CALC = "FFFFFF"    # White for calculations
    COLOR_LINK = "D3D3D3"    # Light gray for links

    # Define border styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ============================================================================
    # INPUTS SHEET
    # ============================================================================

    print("Building Inputs sheet...")

    # Headers
    ws_inputs['A1'] = "FINANCIAL MODEL INPUTS"
    ws_inputs['A1'].font = Font(bold=True, size=14)

    ws_inputs['A3'] = "MODEL ASSUMPTIONS & DRIVERS"
    ws_inputs['A3'].font = Font(bold=True, size=12)

    # Timeline
    ws_inputs['A5'] = "Timeline"
    ws_inputs['A5'].font = Font(bold=True)
    ws_inputs['A6'] = "Number of Periods (Months)"
    ws_inputs['B6'] = NUM_MONTHS
    ws_inputs['B6'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")

    # Revenue Assumptions
    row = 9
    ws_inputs[f'A{row}'] = "REVENUE ASSUMPTIONS"
    ws_inputs[f'A{row}'].font = Font(bold=True, underline="single")
    row += 1

    ws_inputs[f'A{row}'] = "Month 1 Revenue"
    ws_inputs[f'B{row}'] = 100000
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '#,##0'
    row += 1

    ws_inputs[f'A{row}'] = "Monthly Revenue Growth %"
    ws_inputs[f'B{row}'] = 0.05  # 5% growth
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '0.0%'
    row += 1

    # COGS Assumptions
    row += 1
    ws_inputs[f'A{row}'] = "COST OF GOODS SOLD"
    ws_inputs[f'A{row}'].font = Font(bold=True, underline="single")
    row += 1

    ws_inputs[f'A{row}'] = "COGS % of Revenue"
    ws_inputs[f'B{row}'] = 0.40  # 40%
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '0.0%'
    row += 1

    # Operating Expenses
    row += 1
    ws_inputs[f'A{row}'] = "OPERATING EXPENSES"
    ws_inputs[f'A{row}'].font = Font(bold=True, underline="single")
    row += 1

    ws_inputs[f'A{row}'] = "Sales & Marketing (Monthly)"
    ws_inputs[f'B{row}'] = 15000
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '#,##0'
    row += 1

    ws_inputs[f'A{row}'] = "General & Admin (Monthly)"
    ws_inputs[f'B{row}'] = 20000
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '#,##0'
    row += 1

    # Working Capital Assumptions
    row += 1
    ws_inputs[f'A{row}'] = "WORKING CAPITAL ASSUMPTIONS"
    ws_inputs[f'A{row}'].font = Font(bold=True, underline="single")
    row += 1

    ws_inputs[f'A{row}'] = "AR Collection Period (Days)"
    ws_inputs[f'B{row}'] = 45
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '0'
    row += 1

    ws_inputs[f'A{row}'] = "Inventory Days on Hand"
    ws_inputs[f'B{row}'] = 60
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '0'
    row += 1

    ws_inputs[f'A{row}'] = "AP Payment Period (Days)"
    ws_inputs[f'B{row}'] = 30
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '0'
    row += 1

    ws_inputs[f'A{row}'] = "Prepayments (% of Monthly OpEx)"
    ws_inputs[f'B{row}'] = 0.10  # 10%
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '0.0%'
    row += 1

    ws_inputs[f'A{row}'] = "Accrued Expenses (% of Monthly OpEx)"
    ws_inputs[f'B{row}'] = 0.15  # 15%
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '0.0%'
    row += 1

    ws_inputs[f'A{row}'] = "Deferred Revenue (% of Monthly Revenue)"
    ws_inputs[f'B{row}'] = 0.20  # 20%
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '0.0%'
    row += 1

    # Fixed Assets
    row += 1
    ws_inputs[f'A{row}'] = "FIXED ASSETS"
    ws_inputs[f'A{row}'].font = Font(bold=True, underline="single")
    row += 1

    ws_inputs[f'A{row}'] = "Opening PP&E (Gross)"
    ws_inputs[f'B{row}'] = 500000
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '#,##0'
    row += 1

    ws_inputs[f'A{row}'] = "Opening Accumulated Depreciation"
    ws_inputs[f'B{row}'] = 100000
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '#,##0'
    row += 1

    ws_inputs[f'A{row}'] = "Monthly Capex"
    ws_inputs[f'B{row}'] = 10000
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '#,##0'
    row += 1

    ws_inputs[f'A{row}'] = "Depreciation Rate (Annual %)"
    ws_inputs[f'B{row}'] = 0.20  # 20% annual
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '0.0%'
    row += 1

    # Debt
    row += 1
    ws_inputs[f'A{row}'] = "DEBT"
    ws_inputs[f'A{row}'].font = Font(bold=True, underline="single")
    row += 1

    ws_inputs[f'A{row}'] = "Opening Debt Balance"
    ws_inputs[f'B{row}'] = 200000
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '#,##0'
    row += 1

    ws_inputs[f'A{row}'] = "Interest Rate (Annual %)"
    ws_inputs[f'B{row}'] = 0.06  # 6% annual
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '0.0%'
    row += 1

    ws_inputs[f'A{row}'] = "Monthly Principal Repayment"
    ws_inputs[f'B{row}'] = 5000
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '#,##0'
    row += 1

    # Equity
    row += 1
    ws_inputs[f'A{row}'] = "EQUITY"
    ws_inputs[f'A{row}'].font = Font(bold=True, underline="single")
    row += 1

    ws_inputs[f'A{row}'] = "Opening Share Capital"
    ws_inputs[f'B{row}'] = 300000
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '#,##0'
    row += 1

    ws_inputs[f'A{row}'] = "Opening Retained Earnings"
    ws_inputs[f'B{row}'] = 150000
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '#,##0'
    row += 1

    # Cash
    row += 1
    ws_inputs[f'A{row}'] = "CASH"
    ws_inputs[f'A{row}'].font = Font(bold=True, underline="single")
    row += 1

    ws_inputs[f'A{row}'] = "Opening Cash Balance"
    ws_inputs[f'B{row}'] = 200000
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '#,##0'
    row += 1

    # Tax
    row += 1
    ws_inputs[f'A{row}'] = "TAX"
    ws_inputs[f'A{row}'].font = Font(bold=True, underline="single")
    row += 1

    ws_inputs[f'A{row}'] = "Corporate Tax Rate %"
    ws_inputs[f'B{row}'] = 0.25  # 25%
    ws_inputs[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    ws_inputs[f'B{row}'].number_format = '0.0%'

    # Set column widths
    ws_inputs.column_dimensions['A'].width = 35
    ws_inputs.column_dimensions['B'].width = 20

    # ============================================================================
    # WORKINGS SHEET - CONTROL ACCOUNTS
    # ============================================================================

    print("Building Workings sheet with control accounts...")

    # Month headers
    ws_workings['A1'] = "CONTROL ACCOUNTS - WORKINGS"
    ws_workings['A1'].font = Font(bold=True, size=14)

    ws_workings['A3'] = "Period"
    ws_workings['A3'].font = Font(bold=True)
    ws_workings['A3'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    ws_workings['A3'].font = Font(bold=True, color="FFFFFF")

    # Create month headers
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)  # Start from column B
        month_num = (START_MONTH + i - 1) % 12 + 1
        year = START_YEAR + (START_MONTH + i - 1) // 12
        month_name = calendar.month_abbr[month_num]
        ws_workings[f'{col}3'] = f"{month_name}-{year % 100:02d}"
        ws_workings[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws_workings[f'{col}3'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        ws_workings[f'{col}3'].alignment = Alignment(horizontal='center')

    current_row = 5

    # ============================================================================
    # REVENUE & COGS DRIVERS (for control accounts)
    # ============================================================================

    ws_workings[f'A{current_row}'] = "REVENUE & COGS DRIVERS"
    ws_workings[f'A{current_row}'].font = Font(bold=True, size=11, underline="single")
    current_row += 1

    # Revenue
    ws_workings[f'A{current_row}'] = "Revenue (Credit Sales)"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        if i == 0:
            ws_workings[f'{col}{current_row}'] = f"=Inputs!$B$10"
        else:
            prev_col = get_column_letter(i + 1)
            ws_workings[f'{col}{current_row}'] = f"={prev_col}{current_row}*(1+Inputs!$B$11)"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # COGS
    ws_workings[f'A{current_row}'] = "Cost of Goods Sold"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}{current_row-1}*Inputs!$B$15"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Purchases (for inventory)
    ws_workings[f'A{current_row}'] = "Inventory Purchases"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        # Purchases = COGS + (Target Inventory - Current Inventory)
        # Simplified: Purchases to maintain inventory days
        ws_workings[f'{col}{current_row}'] = f"={col}{current_row-1}*1.1"  # Purchases slightly more than COGS
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 2

    # ============================================================================
    # ACCOUNTS RECEIVABLE CONTROL ACCOUNT
    # ============================================================================

    ws_workings[f'A{current_row}'] = "ACCOUNTS RECEIVABLE CONTROL"
    ws_workings[f'A{current_row}'].font = Font(bold=True, size=11, underline="single")
    current_row += 1

    ar_opening_row = current_row
    ws_workings[f'A{current_row}'] = "Opening Balance"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        if i == 0:
            # Calculate opening AR based on collection period
            ws_workings[f'{col}{current_row}'] = f"=Inputs!$B$10*Inputs!$B$21/30"
        else:
            prev_col = get_column_letter(i + 1)
            ws_workings[f'{col}{current_row}'] = f"={prev_col}{current_row+3}"  # Prior month closing
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Credit Sales (increases)
    ar_sales_row = current_row
    ws_workings[f'A{current_row}'] = "  + Credit Sales"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}6"  # Link to Revenue
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Cash Collections (decreases)
    ar_collections_row = current_row
    ws_workings[f'A{current_row}'] = "  - Cash Collections"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        # Collections based on AR balance and collection period
        ws_workings[f'{col}{current_row}'] = f"={col}{ar_opening_row}*30/Inputs!$B$21"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Closing Balance
    ar_closing_row = current_row
    ws_workings[f'A{current_row}'] = "Closing Balance"
    ws_workings[f'A{current_row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}{ar_opening_row}+{col}{ar_sales_row}-{col}{ar_collections_row}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
        ws_workings[f'{col}{current_row}'].font = Font(bold=True)
    current_row += 2

    # ============================================================================
    # INVENTORY CONTROL ACCOUNT
    # ============================================================================

    ws_workings[f'A{current_row}'] = "INVENTORY CONTROL"
    ws_workings[f'A{current_row}'].font = Font(bold=True, size=11, underline="single")
    current_row += 1

    inv_opening_row = current_row
    ws_workings[f'A{current_row}'] = "Opening Balance"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        if i == 0:
            # Opening inventory based on COGS and days on hand
            ws_workings[f'{col}{current_row}'] = f"=Inputs!$B$15*Inputs!$B$10*Inputs!$B$22/30"
        else:
            prev_col = get_column_letter(i + 1)
            ws_workings[f'{col}{current_row}'] = f"={prev_col}{current_row+3}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Purchases (increases)
    inv_purchases_row = current_row
    ws_workings[f'A{current_row}'] = "  + Purchases"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}8"  # Link to Purchases
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # COGS withdrawal (decreases)
    inv_cogs_row = current_row
    ws_workings[f'A{current_row}'] = "  - COGS Withdrawal"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}7"  # Link to COGS
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Closing Balance
    inv_closing_row = current_row
    ws_workings[f'A{current_row}'] = "Closing Balance"
    ws_workings[f'A{current_row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}{inv_opening_row}+{col}{inv_purchases_row}-{col}{inv_cogs_row}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
        ws_workings[f'{col}{current_row}'].font = Font(bold=True)
    current_row += 2

    # ============================================================================
    # PREPAYMENTS CONTROL ACCOUNT
    # ============================================================================

    ws_workings[f'A{current_row}'] = "PREPAYMENTS CONTROL"
    ws_workings[f'A{current_row}'].font = Font(bold=True, size=11, underline="single")
    current_row += 1

    prep_opening_row = current_row
    ws_workings[f'A{current_row}'] = "Opening Balance"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        if i == 0:
            ws_workings[f'{col}{current_row}'] = f"=(Inputs!$B$17+Inputs!$B$18)*Inputs!$B$24"
        else:
            prev_col = get_column_letter(i + 1)
            ws_workings[f'{col}{current_row}'] = f"={prev_col}{current_row+3}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Cash paid in advance (increases)
    prep_additions_row = current_row
    ws_workings[f'A{current_row}'] = "  + Cash Prepayments"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"=(Inputs!$B$17+Inputs!$B$18)*Inputs!$B$24"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Expensed to P&L (decreases)
    prep_expense_row = current_row
    ws_workings[f'A{current_row}'] = "  - Expensed to P&L"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"=(Inputs!$B$17+Inputs!$B$18)*Inputs!$B$24"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Closing Balance
    prep_closing_row = current_row
    ws_workings[f'A{current_row}'] = "Closing Balance"
    ws_workings[f'A{current_row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}{prep_opening_row}+{col}{prep_additions_row}-{col}{prep_expense_row}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
        ws_workings[f'{col}{current_row}'].font = Font(bold=True)
    current_row += 2

    # ============================================================================
    # ACCOUNTS PAYABLE CONTROL ACCOUNT
    # ============================================================================

    ws_workings[f'A{current_row}'] = "ACCOUNTS PAYABLE CONTROL"
    ws_workings[f'A{current_row}'].font = Font(bold=True, size=11, underline="single")
    current_row += 1

    ap_opening_row = current_row
    ws_workings[f'A{current_row}'] = "Opening Balance"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        if i == 0:
            ws_workings[f'{col}{current_row}'] = f"={col}8*Inputs!$B$23/30"  # Based on purchases and payment terms
        else:
            prev_col = get_column_letter(i + 1)
            ws_workings[f'{col}{current_row}'] = f"={prev_col}{current_row+3}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Purchases on credit (increases)
    ap_purchases_row = current_row
    ws_workings[f'A{current_row}'] = "  + Credit Purchases"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}8"  # Link to Purchases
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Cash payments (decreases)
    ap_payments_row = current_row
    ws_workings[f'A{current_row}'] = "  - Cash Payments"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}{ap_opening_row}*30/Inputs!$B$23"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Closing Balance
    ap_closing_row = current_row
    ws_workings[f'A{current_row}'] = "Closing Balance"
    ws_workings[f'A{current_row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}{ap_opening_row}+{col}{ap_purchases_row}-{col}{ap_payments_row}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
        ws_workings[f'{col}{current_row}'].font = Font(bold=True)
    current_row += 2

    # ============================================================================
    # ACCRUED EXPENSES CONTROL ACCOUNT
    # ============================================================================

    ws_workings[f'A{current_row}'] = "ACCRUED EXPENSES CONTROL"
    ws_workings[f'A{current_row}'].font = Font(bold=True, size=11, underline="single")
    current_row += 1

    accrued_opening_row = current_row
    ws_workings[f'A{current_row}'] = "Opening Balance"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        if i == 0:
            ws_workings[f'{col}{current_row}'] = f"=(Inputs!$B$17+Inputs!$B$18)*Inputs!$B$25"
        else:
            prev_col = get_column_letter(i + 1)
            ws_workings[f'{col}{current_row}'] = f"={prev_col}{current_row+3}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Expenses accrued (increases)
    accrued_additions_row = current_row
    ws_workings[f'A{current_row}'] = "  + Expenses Accrued"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"=(Inputs!$B$17+Inputs!$B$18)*Inputs!$B$25"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Cash payments (decreases)
    accrued_payments_row = current_row
    ws_workings[f'A{current_row}'] = "  - Cash Payments"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"=(Inputs!$B$17+Inputs!$B$18)*Inputs!$B$25"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Closing Balance
    accrued_closing_row = current_row
    ws_workings[f'A{current_row}'] = "Closing Balance"
    ws_workings[f'A{current_row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}{accrued_opening_row}+{col}{accrued_additions_row}-{col}{accrued_payments_row}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
        ws_workings[f'{col}{current_row}'].font = Font(bold=True)
    current_row += 2

    # ============================================================================
    # DEFERRED REVENUE CONTROL ACCOUNT
    # ============================================================================

    ws_workings[f'A{current_row}'] = "DEFERRED REVENUE CONTROL"
    ws_workings[f'A{current_row}'].font = Font(bold=True, size=11, underline="single")
    current_row += 1

    deferred_opening_row = current_row
    ws_workings[f'A{current_row}'] = "Opening Balance"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        if i == 0:
            ws_workings[f'{col}{current_row}'] = f"=Inputs!$B$10*Inputs!$B$26"
        else:
            prev_col = get_column_letter(i + 1)
            ws_workings[f'{col}{current_row}'] = f"={prev_col}{current_row+3}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Cash received in advance (increases)
    deferred_additions_row = current_row
    ws_workings[f'A{current_row}'] = "  + Cash Received in Advance"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}6*Inputs!$B$26"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Revenue recognized (decreases)
    deferred_recognition_row = current_row
    ws_workings[f'A{current_row}'] = "  - Revenue Recognized"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}6*Inputs!$B$26"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Closing Balance
    deferred_closing_row = current_row
    ws_workings[f'A{current_row}'] = "Closing Balance"
    ws_workings[f'A{current_row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}{deferred_opening_row}+{col}{deferred_additions_row}-{col}{deferred_recognition_row}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
        ws_workings[f'{col}{current_row}'].font = Font(bold=True)
    current_row += 2

    # ============================================================================
    # FIXED ASSETS CONTROL ACCOUNT (PP&E GROSS)
    # ============================================================================

    ws_workings[f'A{current_row}'] = "PP&E GROSS CONTROL"
    ws_workings[f'A{current_row}'].font = Font(bold=True, size=11, underline="single")
    current_row += 1

    ppe_gross_opening_row = current_row
    ws_workings[f'A{current_row}'] = "Opening Balance"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        if i == 0:
            ws_workings[f'{col}{current_row}'] = f"=Inputs!$B$30"
        else:
            prev_col = get_column_letter(i + 1)
            ws_workings[f'{col}{current_row}'] = f"={prev_col}{current_row+3}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Capex additions (increases)
    ppe_capex_row = current_row
    ws_workings[f'A{current_row}'] = "  + Capital Expenditure"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"=Inputs!$B$32"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Disposals (decreases)
    ppe_disposals_row = current_row
    ws_workings[f'A{current_row}'] = "  - Disposals"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = 0
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Closing Balance
    ppe_gross_closing_row = current_row
    ws_workings[f'A{current_row}'] = "Closing Balance"
    ws_workings[f'A{current_row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}{ppe_gross_opening_row}+{col}{ppe_capex_row}-{col}{ppe_disposals_row}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
        ws_workings[f'{col}{current_row}'].font = Font(bold=True)
    current_row += 2

    # ============================================================================
    # ACCUMULATED DEPRECIATION CONTROL ACCOUNT
    # ============================================================================

    ws_workings[f'A{current_row}'] = "ACCUMULATED DEPRECIATION CONTROL"
    ws_workings[f'A{current_row}'].font = Font(bold=True, size=11, underline="single")
    current_row += 1

    acc_dep_opening_row = current_row
    ws_workings[f'A{current_row}'] = "Opening Balance"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        if i == 0:
            ws_workings[f'{col}{current_row}'] = f"=Inputs!$B$31"
        else:
            prev_col = get_column_letter(i + 1)
            ws_workings[f'{col}{current_row}'] = f"={prev_col}{current_row+3}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Depreciation expense (increases)
    acc_dep_expense_row = current_row
    ws_workings[f'A{current_row}'] = "  + Depreciation Expense"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        # Monthly depreciation = Annual rate / 12 * Gross PPE
        ws_workings[f'{col}{current_row}'] = f"={col}{ppe_gross_opening_row}*Inputs!$B$33/12"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Disposals (decreases accumulated depreciation)
    acc_dep_disposals_row = current_row
    ws_workings[f'A{current_row}'] = "  - Disposals"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = 0
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Closing Balance
    acc_dep_closing_row = current_row
    ws_workings[f'A{current_row}'] = "Closing Balance"
    ws_workings[f'A{current_row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}{acc_dep_opening_row}+{col}{acc_dep_expense_row}-{col}{acc_dep_disposals_row}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
        ws_workings[f'{col}{current_row}'].font = Font(bold=True)
    current_row += 2

    # ============================================================================
    # DEBT CONTROL ACCOUNT
    # ============================================================================

    ws_workings[f'A{current_row}'] = "DEBT CONTROL"
    ws_workings[f'A{current_row}'].font = Font(bold=True, size=11, underline="single")
    current_row += 1

    debt_opening_row = current_row
    ws_workings[f'A{current_row}'] = "Opening Balance"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        if i == 0:
            ws_workings[f'{col}{current_row}'] = f"=Inputs!$B$37"
        else:
            prev_col = get_column_letter(i + 1)
            ws_workings[f'{col}{current_row}'] = f"={prev_col}{current_row+3}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Drawdowns (increases)
    debt_drawdowns_row = current_row
    ws_workings[f'A{current_row}'] = "  + Drawdowns"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = 0
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Principal repayments (decreases)
    debt_repayments_row = current_row
    ws_workings[f'A{current_row}'] = "  - Principal Repayments"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"=Inputs!$B$39"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Closing Balance
    debt_closing_row = current_row
    ws_workings[f'A{current_row}'] = "Closing Balance"
    ws_workings[f'A{current_row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}{debt_opening_row}+{col}{debt_drawdowns_row}-{col}{debt_repayments_row}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
        ws_workings[f'{col}{current_row}'].font = Font(bold=True)
    current_row += 2

    # Interest calculation (for P&L)
    debt_interest_row = current_row
    ws_workings[f'A{current_row}'] = "Interest Expense (Monthly)"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}{debt_opening_row}*Inputs!$B$38/12"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 2

    # ============================================================================
    # EQUITY CONTROL ACCOUNTS
    # ============================================================================

    ws_workings[f'A{current_row}'] = "SHARE CAPITAL CONTROL"
    ws_workings[f'A{current_row}'].font = Font(bold=True, size=11, underline="single")
    current_row += 1

    equity_opening_row = current_row
    ws_workings[f'A{current_row}'] = "Opening Balance"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        if i == 0:
            ws_workings[f'{col}{current_row}'] = f"=Inputs!$B$43"
        else:
            prev_col = get_column_letter(i + 1)
            ws_workings[f'{col}{current_row}'] = f"={prev_col}{current_row+3}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # New issuances (increases)
    equity_issuances_row = current_row
    ws_workings[f'A{current_row}'] = "  + New Issuances"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = 0
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Buybacks (decreases)
    equity_buybacks_row = current_row
    ws_workings[f'A{current_row}'] = "  - Buybacks"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = 0
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Closing Balance
    equity_closing_row = current_row
    ws_workings[f'A{current_row}'] = "Closing Balance"
    ws_workings[f'A{current_row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}{equity_opening_row}+{col}{equity_issuances_row}-{col}{equity_buybacks_row}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
        ws_workings[f'{col}{current_row}'].font = Font(bold=True)
    current_row += 2

    # ============================================================================
    # RETAINED EARNINGS CONTROL
    # ============================================================================

    ws_workings[f'A{current_row}'] = "RETAINED EARNINGS CONTROL"
    ws_workings[f'A{current_row}'].font = Font(bold=True, size=11, underline="single")
    current_row += 1

    re_opening_row = current_row
    ws_workings[f'A{current_row}'] = "Opening Balance"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        if i == 0:
            ws_workings[f'{col}{current_row}'] = f"=Inputs!$B$44"
        else:
            prev_col = get_column_letter(i + 1)
            ws_workings[f'{col}{current_row}'] = f"={prev_col}{current_row+3}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Net Income (increases)
    re_netincome_row = current_row
    ws_workings[f'A{current_row}'] = "  + Net Income"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"='Income Statement'!{col}19"  # Will link to Net Income
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Dividends (decreases)
    re_dividends_row = current_row
    ws_workings[f'A{current_row}'] = "  - Dividends"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = 0
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Closing Balance
    re_closing_row = current_row
    ws_workings[f'A{current_row}'] = "Closing Balance"
    ws_workings[f'A{current_row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"={col}{re_opening_row}+{col}{re_netincome_row}-{col}{re_dividends_row}"
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
        ws_workings[f'{col}{current_row}'].font = Font(bold=True)
    current_row += 2

    # ============================================================================
    # CASH CONTROL ACCOUNT
    # ============================================================================

    ws_workings[f'A{current_row}'] = "CASH CONTROL"
    ws_workings[f'A{current_row}'].font = Font(bold=True, size=11, underline="single")
    current_row += 1

    cash_opening_row = current_row
    ws_workings[f'A{current_row}'] = "Opening Balance"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        if i == 0:
            ws_workings[f'{col}{current_row}'] = f"=Inputs!$B$47"
        else:
            prev_col = get_column_letter(i + 1)
            ws_workings[f'{col}{current_row}'] = f"={prev_col}{current_row+1}"  # Prior closing
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
    current_row += 1

    # Closing Balance (from Cash Flow Statement)
    cash_closing_row = current_row
    ws_workings[f'A{current_row}'] = "Closing Balance"
    ws_workings[f'A{current_row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings[f'{col}{current_row}'] = f"='Cash Flow Statement'!{col}31"  # Will link to closing cash
        ws_workings[f'{col}{current_row}'].number_format = '#,##0'
        ws_workings[f'{col}{current_row}'].font = Font(bold=True)
    current_row += 2

    # Set column widths for Workings
    ws_workings.column_dimensions['A'].width = 35
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_workings.column_dimensions[col].width = 12

    # Store row references for later use
    control_rows = {
        'ar_closing': ar_closing_row,
        'inv_closing': inv_closing_row,
        'prep_closing': prep_closing_row,
        'ap_closing': ap_closing_row,
        'accrued_closing': accrued_closing_row,
        'deferred_closing': deferred_closing_row,
        'ppe_gross_closing': ppe_gross_closing_row,
        'acc_dep_closing': acc_dep_closing_row,
        'debt_closing': debt_closing_row,
        'equity_closing': equity_closing_row,
        're_closing': re_closing_row,
        'cash_closing': cash_closing_row,
        'cash_opening': cash_opening_row,
        'debt_interest': debt_interest_row,
        'acc_dep_expense': acc_dep_expense_row,
        'ar_collections': ar_collections_row,
        'ap_payments': ap_payments_row,
        'ppe_capex': ppe_capex_row,
        'debt_repayments': debt_repayments_row,
        'prep_additions': prep_additions_row,
        'accrued_payments': accrued_payments_row,
        'deferred_additions': deferred_additions_row
    }

    # ============================================================================
    # INCOME STATEMENT
    # ============================================================================

    print("Building Income Statement...")

    ws_income['A1'] = "INCOME STATEMENT"
    ws_income['A1'].font = Font(bold=True, size=14)

    ws_income['A3'] = "Period"
    ws_income['A3'].font = Font(bold=True, color="FFFFFF")
    ws_income['A3'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

    # Month headers
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        month_num = (START_MONTH + i - 1) % 12 + 1
        year = START_YEAR + (START_MONTH + i - 1) // 12
        month_name = calendar.month_abbr[month_num]
        ws_income[f'{col}3'] = f"{month_name}-{year % 100:02d}"
        ws_income[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws_income[f'{col}3'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        ws_income[f'{col}3'].alignment = Alignment(horizontal='center')

    row = 5

    # Revenue
    ws_income[f'A{row}'] = "Revenue"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_income[f'{col}{row}'] = f"=Workings!{col}6"
        ws_income[f'{col}{row}'].number_format = '#,##0'
    row += 1

    # COGS
    ws_income[f'A{row}'] = "Cost of Goods Sold"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_income[f'{col}{row}'] = f"=-Workings!{col}7"
        ws_income[f'{col}{row}'].number_format = '#,##0'
    row += 1

    # Gross Profit
    ws_income[f'A{row}'] = "Gross Profit"
    ws_income[f'A{row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_income[f'{col}{row}'] = f"={col}{row-2}+{col}{row-1}"
        ws_income[f'{col}{row}'].number_format = '#,##0'
        ws_income[f'{col}{row}'].font = Font(bold=True)
    row += 2

    # Operating Expenses
    ws_income[f'A{row}'] = "Operating Expenses:"
    ws_income[f'A{row}'].font = Font(bold=True)
    row += 1

    ws_income[f'A{row}'] = "  Sales & Marketing"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_income[f'{col}{row}'] = f"=-Inputs!$B$17"
        ws_income[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_income[f'A{row}'] = "  General & Administrative"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_income[f'{col}{row}'] = f"=-Inputs!$B$18"
        ws_income[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_income[f'A{row}'] = "  Depreciation"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_income[f'{col}{row}'] = f"=-Workings!{col}{control_rows['acc_dep_expense']}"
        ws_income[f'{col}{row}'].number_format = '#,##0'
    row += 1

    # EBIT
    ws_income[f'A{row}'] = "EBIT"
    ws_income[f'A{row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_income[f'{col}{row}'] = f"={col}7+{col}10+{col}11+{col}12"
        ws_income[f'{col}{row}'].number_format = '#,##0'
        ws_income[f'{col}{row}'].font = Font(bold=True)
    row += 2

    # Interest Expense
    ws_income[f'A{row}'] = "Interest Expense"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_income[f'{col}{row}'] = f"=-Workings!{col}{control_rows['debt_interest']}"
        ws_income[f'{col}{row}'].number_format = '#,##0'
    row += 1

    # EBT
    ws_income[f'A{row}'] = "Earnings Before Tax"
    ws_income[f'A{row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_income[f'{col}{row}'] = f"={col}{row-2}+{col}{row-1}"
        ws_income[f'{col}{row}'].number_format = '#,##0'
        ws_income[f'{col}{row}'].font = Font(bold=True)
    row += 2

    # Tax
    ws_income[f'A{row}'] = "Income Tax"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_income[f'{col}{row}'] = f"=-MAX({col}{row-1},0)*Inputs!$B$50"
        ws_income[f'{col}{row}'].number_format = '#,##0'
    row += 1

    # Net Income
    ws_income[f'A{row}'] = "Net Income"
    ws_income[f'A{row}'].font = Font(bold=True, size=11)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_income[f'{col}{row}'] = f"={col}{row-2}+{col}{row-1}"
        ws_income[f'{col}{row}'].number_format = '#,##0'
        ws_income[f'{col}{row}'].font = Font(bold=True)
        ws_income[f'{col}{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # Set column widths
    ws_income.column_dimensions['A'].width = 30
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_income.column_dimensions[col].width = 12

    # ============================================================================
    # BALANCE SHEET
    # ============================================================================

    print("Building Balance Sheet...")

    ws_balance['A1'] = "BALANCE SHEET"
    ws_balance['A1'].font = Font(bold=True, size=14)

    ws_balance['A3'] = "Period"
    ws_balance['A3'].font = Font(bold=True, color="FFFFFF")
    ws_balance['A3'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

    # Month headers
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        month_num = (START_MONTH + i - 1) % 12 + 1
        year = START_YEAR + (START_MONTH + i - 1) // 12
        month_name = calendar.month_abbr[month_num]
        ws_balance[f'{col}3'] = f"{month_name}-{year % 100:02d}"
        ws_balance[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws_balance[f'{col}3'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        ws_balance[f'{col}3'].alignment = Alignment(horizontal='center')

    row = 5

    # ASSETS
    ws_balance[f'A{row}'] = "ASSETS"
    ws_balance[f'A{row}'].font = Font(bold=True, size=12, underline="single")
    row += 1

    # Current Assets
    ws_balance[f'A{row}'] = "Current Assets:"
    ws_balance[f'A{row}'].font = Font(bold=True)
    row += 1

    ws_balance[f'A{row}'] = "  Cash"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"=Workings!{col}{control_rows['cash_closing']}"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_balance[f'A{row}'] = "  Accounts Receivable"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"=Workings!{col}{control_rows['ar_closing']}"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_balance[f'A{row}'] = "  Inventory"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"=Workings!{col}{control_rows['inv_closing']}"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_balance[f'A{row}'] = "  Prepayments"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"=Workings!{col}{control_rows['prep_closing']}"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_balance[f'A{row}'] = "Total Current Assets"
    ws_balance[f'A{row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"=SUM({col}7:{col}10)"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
        ws_balance[f'{col}{row}'].font = Font(bold=True)
    row += 2

    # Fixed Assets
    ws_balance[f'A{row}'] = "Fixed Assets:"
    ws_balance[f'A{row}'].font = Font(bold=True)
    row += 1

    ws_balance[f'A{row}'] = "  PP&E (Gross)"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"=Workings!{col}{control_rows['ppe_gross_closing']}"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_balance[f'A{row}'] = "  Less: Accumulated Depreciation"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"=-Workings!{col}{control_rows['acc_dep_closing']}"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_balance[f'A{row}'] = "Net PP&E"
    ws_balance[f'A{row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"={col}{row-2}+{col}{row-1}"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
        ws_balance[f'{col}{row}'].font = Font(bold=True)
    row += 2

    # Total Assets
    ws_balance[f'A{row}'] = "TOTAL ASSETS"
    ws_balance[f'A{row}'].font = Font(bold=True, size=11)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"={col}11+{col}16"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
        ws_balance[f'{col}{row}'].font = Font(bold=True)
        ws_balance[f'{col}{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    row += 2

    # LIABILITIES
    ws_balance[f'A{row}'] = "LIABILITIES"
    ws_balance[f'A{row}'].font = Font(bold=True, size=12, underline="single")
    row += 1

    # Current Liabilities
    ws_balance[f'A{row}'] = "Current Liabilities:"
    ws_balance[f'A{row}'].font = Font(bold=True)
    row += 1

    ws_balance[f'A{row}'] = "  Accounts Payable"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"=Workings!{col}{control_rows['ap_closing']}"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_balance[f'A{row}'] = "  Accrued Expenses"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"=Workings!{col}{control_rows['accrued_closing']}"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_balance[f'A{row}'] = "  Deferred Revenue"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"=Workings!{col}{control_rows['deferred_closing']}"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_balance[f'A{row}'] = "Total Current Liabilities"
    ws_balance[f'A{row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"=SUM({col}22:{col}24)"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
        ws_balance[f'{col}{row}'].font = Font(bold=True)
    row += 2

    # Long-term Liabilities
    ws_balance[f'A{row}'] = "Long-term Liabilities:"
    ws_balance[f'A{row}'].font = Font(bold=True)
    row += 1

    ws_balance[f'A{row}'] = "  Long-term Debt"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"=Workings!{col}{control_rows['debt_closing']}"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_balance[f'A{row}'] = "Total Long-term Liabilities"
    ws_balance[f'A{row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"={col}{row-1}"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
        ws_balance[f'{col}{row}'].font = Font(bold=True)
    row += 2

    # Total Liabilities
    ws_balance[f'A{row}'] = "TOTAL LIABILITIES"
    ws_balance[f'A{row}'].font = Font(bold=True, size=11)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"={col}25+{col}29"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
        ws_balance[f'{col}{row}'].font = Font(bold=True)
        ws_balance[f'{col}{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    row += 2

    # EQUITY
    ws_balance[f'A{row}'] = "EQUITY"
    ws_balance[f'A{row}'].font = Font(bold=True, size=12, underline="single")
    row += 1

    ws_balance[f'A{row}'] = "  Share Capital"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"=Workings!{col}{control_rows['equity_closing']}"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_balance[f'A{row}'] = "  Retained Earnings"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"=Workings!{col}{control_rows['re_closing']}"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_balance[f'A{row}'] = "TOTAL EQUITY"
    ws_balance[f'A{row}'].font = Font(bold=True, size=11)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"=SUM({col}{row-2}:{col}{row-1})"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
        ws_balance[f'{col}{row}'].font = Font(bold=True)
        ws_balance[f'{col}{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    row += 2

    # Total Liabilities & Equity
    ws_balance[f'A{row}'] = "TOTAL LIABILITIES & EQUITY"
    ws_balance[f'A{row}'].font = Font(bold=True, size=11)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"={col}31+{col}36"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
        ws_balance[f'{col}{row}'].font = Font(bold=True)
        ws_balance[f'{col}{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    row += 2

    # Balance Check
    ws_balance[f'A{row}'] = "BALANCE CHECK (Should be 0)"
    ws_balance[f'A{row}'].font = Font(bold=True, color="FF0000")
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance[f'{col}{row}'] = f"={col}18-{col}38"
        ws_balance[f'{col}{row}'].number_format = '#,##0'
        ws_balance[f'{col}{row}'].font = Font(bold=True, color="FF0000")

    # Set column widths
    ws_balance.column_dimensions['A'].width = 35
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_balance.column_dimensions[col].width = 12

    # ============================================================================
    # CASH FLOW STATEMENT (Indirect Method)
    # ============================================================================

    print("Building Cash Flow Statement...")

    ws_cashflow['A1'] = "CASH FLOW STATEMENT (Indirect Method)"
    ws_cashflow['A1'].font = Font(bold=True, size=14)

    ws_cashflow['A3'] = "Period"
    ws_cashflow['A3'].font = Font(bold=True, color="FFFFFF")
    ws_cashflow['A3'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

    # Month headers
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        month_num = (START_MONTH + i - 1) % 12 + 1
        year = START_YEAR + (START_MONTH + i - 1) // 12
        month_name = calendar.month_abbr[month_num]
        ws_cashflow[f'{col}3'] = f"{month_name}-{year % 100:02d}"
        ws_cashflow[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws_cashflow[f'{col}3'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        ws_cashflow[f'{col}3'].alignment = Alignment(horizontal='center')

    row = 5

    # Operating Activities
    ws_cashflow[f'A{row}'] = "OPERATING ACTIVITIES"
    ws_cashflow[f'A{row}'].font = Font(bold=True, size=11, underline="single")
    row += 1

    ws_cashflow[f'A{row}'] = "Net Income"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_cashflow[f'{col}{row}'] = f"='Income Statement'!{col}19"
        ws_cashflow[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_cashflow[f'A{row}'] = "Adjustments:"
    ws_cashflow[f'A{row}'].font = Font(italic=True)
    row += 1

    ws_cashflow[f'A{row}'] = "  + Depreciation"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_cashflow[f'{col}{row}'] = f"=Workings!{col}{control_rows['acc_dep_expense']}"
        ws_cashflow[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_cashflow[f'A{row}'] = "Changes in Working Capital:"
    ws_cashflow[f'A{row}'].font = Font(italic=True)
    row += 1

    ws_cashflow[f'A{row}'] = "  - Increase in AR"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        if i == 0:
            ws_cashflow[f'{col}{row}'] = f"=-(Workings!{col}{control_rows['ar_closing']}-Workings!B{control_rows['ar_closing']-3})"
        else:
            prev_col = get_column_letter(i + 1)
            ws_cashflow[f'{col}{row}'] = f"=-(Workings!{col}{control_rows['ar_closing']}-Workings!{prev_col}{control_rows['ar_closing']})"
        ws_cashflow[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_cashflow[f'A{row}'] = "  - Increase in Inventory"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        if i == 0:
            ws_cashflow[f'{col}{row}'] = f"=-(Workings!{col}{control_rows['inv_closing']}-Workings!B{control_rows['inv_closing']-3})"
        else:
            prev_col = get_column_letter(i + 1)
            ws_cashflow[f'{col}{row}'] = f"=-(Workings!{col}{control_rows['inv_closing']}-Workings!{prev_col}{control_rows['inv_closing']})"
        ws_cashflow[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_cashflow[f'A{row}'] = "  - Increase in Prepayments"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        if i == 0:
            ws_cashflow[f'{col}{row}'] = f"=-(Workings!{col}{control_rows['prep_closing']}-Workings!B{control_rows['prep_closing']-3})"
        else:
            prev_col = get_column_letter(i + 1)
            ws_cashflow[f'{col}{row}'] = f"=-(Workings!{col}{control_rows['prep_closing']}-Workings!{prev_col}{control_rows['prep_closing']})"
        ws_cashflow[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_cashflow[f'A{row}'] = "  + Increase in AP"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        if i == 0:
            ws_cashflow[f'{col}{row}'] = f"=Workings!{col}{control_rows['ap_closing']}-Workings!B{control_rows['ap_closing']-3}"
        else:
            prev_col = get_column_letter(i + 1)
            ws_cashflow[f'{col}{row}'] = f"=Workings!{col}{control_rows['ap_closing']}-Workings!{prev_col}{control_rows['ap_closing']}"
        ws_cashflow[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_cashflow[f'A{row}'] = "  + Increase in Accrued Expenses"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        if i == 0:
            ws_cashflow[f'{col}{row}'] = f"=Workings!{col}{control_rows['accrued_closing']}-Workings!B{control_rows['accrued_closing']-3}"
        else:
            prev_col = get_column_letter(i + 1)
            ws_cashflow[f'{col}{row}'] = f"=Workings!{col}{control_rows['accrued_closing']}-Workings!{prev_col}{control_rows['accrued_closing']}"
        ws_cashflow[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_cashflow[f'A{row}'] = "  + Increase in Deferred Revenue"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        if i == 0:
            ws_cashflow[f'{col}{row}'] = f"=Workings!{col}{control_rows['deferred_closing']}-Workings!B{control_rows['deferred_closing']-3}"
        else:
            prev_col = get_column_letter(i + 1)
            ws_cashflow[f'{col}{row}'] = f"=Workings!{col}{control_rows['deferred_closing']}-Workings!{prev_col}{control_rows['deferred_closing']}"
        ws_cashflow[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_cashflow[f'A{row}'] = "Cash from Operating Activities"
    ws_cashflow[f'A{row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_cashflow[f'{col}{row}'] = f"={col}6+{col}8+SUM({col}10:{col}15)"
        ws_cashflow[f'{col}{row}'].number_format = '#,##0'
        ws_cashflow[f'{col}{row}'].font = Font(bold=True)
    row += 2

    # Investing Activities
    ws_cashflow[f'A{row}'] = "INVESTING ACTIVITIES"
    ws_cashflow[f'A{row}'].font = Font(bold=True, size=11, underline="single")
    row += 1

    ws_cashflow[f'A{row}'] = "  Capital Expenditure"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_cashflow[f'{col}{row}'] = f"=-Workings!{col}{control_rows['ppe_capex']}"
        ws_cashflow[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_cashflow[f'A{row}'] = "Cash from Investing Activities"
    ws_cashflow[f'A{row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_cashflow[f'{col}{row}'] = f"={col}{row-1}"
        ws_cashflow[f'{col}{row}'].number_format = '#,##0'
        ws_cashflow[f'{col}{row}'].font = Font(bold=True)
    row += 2

    # Financing Activities
    ws_cashflow[f'A{row}'] = "FINANCING ACTIVITIES"
    ws_cashflow[f'A{row}'].font = Font(bold=True, size=11, underline="single")
    row += 1

    ws_cashflow[f'A{row}'] = "  Debt Repayments"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_cashflow[f'{col}{row}'] = f"=-Workings!{col}{control_rows['debt_repayments']}"
        ws_cashflow[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_cashflow[f'A{row}'] = "  Dividends Paid"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_cashflow[f'{col}{row}'] = 0
        ws_cashflow[f'{col}{row}'].number_format = '#,##0'
    row += 1

    ws_cashflow[f'A{row}'] = "Cash from Financing Activities"
    ws_cashflow[f'A{row}'].font = Font(bold=True)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_cashflow[f'{col}{row}'] = f"=SUM({col}{row-2}:{col}{row-1})"
        ws_cashflow[f'{col}{row}'].number_format = '#,##0'
        ws_cashflow[f'{col}{row}'].font = Font(bold=True)
    row += 2

    # Net change in cash
    ws_cashflow[f'A{row}'] = "Net Change in Cash"
    ws_cashflow[f'A{row}'].font = Font(bold=True, size=11)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_cashflow[f'{col}{row}'] = f"={col}16+{col}20+{col}25"
        ws_cashflow[f'{col}{row}'].number_format = '#,##0'
        ws_cashflow[f'{col}{row}'].font = Font(bold=True)
        ws_cashflow[f'{col}{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    row += 2

    # Opening Cash
    ws_cashflow[f'A{row}'] = "Opening Cash"
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_cashflow[f'{col}{row}'] = f"=Workings!{col}{control_rows['cash_opening']}"
        ws_cashflow[f'{col}{row}'].number_format = '#,##0'
    row += 1

    # Closing Cash
    ws_cashflow[f'A{row}'] = "Closing Cash"
    ws_cashflow[f'A{row}'].font = Font(bold=True, size=11)
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_cashflow[f'{col}{row}'] = f"={col}{row-1}+{col}{row-3}"
        ws_cashflow[f'{col}{row}'].number_format = '#,##0'
        ws_cashflow[f'{col}{row}'].font = Font(bold=True)
        ws_cashflow[f'{col}{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    row += 2

    # Check vs Balance Sheet
    ws_cashflow[f'A{row}'] = "Check vs Balance Sheet (Should be 0)"
    ws_cashflow[f'A{row}'].font = Font(bold=True, color="FF0000")
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_cashflow[f'{col}{row}'] = f"={col}{row-2}-'Balance Sheet'!{col}7"
        ws_cashflow[f'{col}{row}'].number_format = '#,##0'
        ws_cashflow[f'{col}{row}'].font = Font(bold=True, color="FF0000")

    # Set column widths
    ws_cashflow.column_dimensions['A'].width = 35
    for i in range(NUM_MONTHS):
        col = get_column_letter(i + 2)
        ws_cashflow.column_dimensions[col].width = 12

    # Save workbook
    filename = "Financial_Model_3_Statement.xlsx"
    wb.save(filename)
    print(f"Financial model saved to {filename}")

    return filename

if __name__ == "__main__":
    print("Creating 3-Statement Financial Model with Control Accounts...")
    create_financial_model()
    print("Model creation complete!")
