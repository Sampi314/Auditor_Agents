"""
5-Year Strategic Financial Model
=================================
Three-Statement Model: Income Statement, Balance Sheet, Cash Flow Statement
Includes full audit module with cross-statement validation.

Author: Claude AI
Date: 2026-02-10
"""

import json
from dataclasses import dataclass, field
from typing import Dict, List, Optional
from copy import deepcopy

# ==============================================================================
# SECTION 1: ASSUMPTIONS & DRIVERS
# ==============================================================================

PROJECTION_YEARS = 5
BASE_YEAR = 2025
YEARS = [BASE_YEAR + i for i in range(PROJECTION_YEARS + 1)]  # 2025 (base) + 5 projected
PROJECTED_YEARS = YEARS[1:]  # 2026-2030


@dataclass
class RevenueAssumptions:
    """Revenue growth drivers."""
    base_revenue: float = 10_000_000.0          # Year 0 revenue
    growth_rates: Dict[int, float] = field(default_factory=lambda: {
        2026: 0.12,   # 12% growth - market expansion phase
        2027: 0.15,   # 15% growth - scaling phase
        2028: 0.18,   # 18% growth - peak growth
        2029: 0.14,   # 14% growth - maturation begins
        2030: 0.10,   # 10% growth - steady state
    })
    # Revenue mix breakdown
    product_revenue_pct: float = 0.65
    service_revenue_pct: float = 0.25
    subscription_revenue_pct: float = 0.10


@dataclass
class CostAssumptions:
    """Cost structure drivers."""
    cogs_pct_of_revenue: Dict[int, float] = field(default_factory=lambda: {
        2025: 0.42, 2026: 0.41, 2027: 0.40, 2028: 0.39,
        2029: 0.38, 2030: 0.37,  # Improving margins via scale
    })
    sga_pct_of_revenue: Dict[int, float] = field(default_factory=lambda: {
        2025: 0.22, 2026: 0.21, 2027: 0.20, 2028: 0.19,
        2029: 0.18, 2030: 0.18,  # Operating leverage
    })
    rd_pct_of_revenue: Dict[int, float] = field(default_factory=lambda: {
        2025: 0.08, 2026: 0.09, 2027: 0.10, 2028: 0.10,
        2029: 0.09, 2030: 0.09,  # Strategic R&D investment
    })
    depreciation_pct_of_ppe: float = 0.10       # Straight-line ~10yr life
    amortization_pct_of_intangibles: float = 0.20  # ~5yr life


@dataclass
class BalanceSheetAssumptions:
    """Balance sheet drivers."""
    # Working capital days
    days_sales_outstanding: float = 45.0
    days_inventory_outstanding: float = 60.0
    days_payable_outstanding: float = 40.0
    prepaid_pct_of_revenue: float = 0.02
    accrued_liabilities_pct_of_opex: float = 0.08

    # Capital structure
    capex_pct_of_revenue: Dict[int, float] = field(default_factory=lambda: {
        2026: 0.06, 2027: 0.07, 2028: 0.08,
        2029: 0.07, 2030: 0.06,
    })
    intangible_investment_pct_of_revenue: Dict[int, float] = field(default_factory=lambda: {
        2026: 0.03, 2027: 0.04, 2028: 0.04,
        2029: 0.03, 2030: 0.03,
    })

    # Base year balance sheet
    base_cash: float = 2_500_000.0
    base_accounts_receivable: float = 1_250_000.0
    base_inventory: float = 1_150_000.0
    base_prepaid_expenses: float = 200_000.0
    base_ppe_gross: float = 4_000_000.0
    base_accumulated_depreciation: float = 1_200_000.0
    base_intangibles_gross: float = 2_000_000.0
    base_accumulated_amortization: float = 800_000.0
    base_other_long_term_assets: float = 500_000.0

    base_accounts_payable: float = 700_000.0
    base_accrued_liabilities: float = 450_000.0
    base_short_term_debt: float = 500_000.0
    base_long_term_debt: float = 3_000_000.0
    base_deferred_revenue: float = 300_000.0
    base_other_long_term_liabilities: float = 400_000.0

    # Equity
    base_common_stock: float = 1_000_000.0
    base_additional_paid_in_capital: float = 2_000_000.0
    base_retained_earnings: float = 1_250_000.0

    # Debt schedule
    interest_rate: float = 0.055                # 5.5% blended rate
    debt_repayment_per_year: float = 300_000.0  # Annual principal repayment
    new_debt_issuance: Dict[int, float] = field(default_factory=lambda: {
        2026: 0.0, 2027: 500_000.0, 2028: 0.0,
        2029: 0.0, 2030: 0.0,
    })

    # Tax & dividends
    tax_rate: float = 0.25
    dividend_payout_ratio: float = 0.20         # 20% of net income


@dataclass
class Assumptions:
    revenue: RevenueAssumptions = field(default_factory=RevenueAssumptions)
    cost: CostAssumptions = field(default_factory=CostAssumptions)
    balance_sheet: BalanceSheetAssumptions = field(default_factory=BalanceSheetAssumptions)


# ==============================================================================
# SECTION 2: INCOME STATEMENT
# ==============================================================================

class IncomeStatement:
    """Builds the projected Income Statement."""

    def __init__(self, assumptions: Assumptions):
        self.a = assumptions
        self.data: Dict[int, Dict[str, float]] = {}

    def build(self, depreciation_schedule: Dict[int, float],
              amortization_schedule: Dict[int, float],
              interest_expense_schedule: Dict[int, float]) -> Dict[int, Dict[str, float]]:
        revenue = self.a.revenue.base_revenue

        for year in YEARS:
            row = {}

            if year == BASE_YEAR:
                row["revenue"] = revenue
            else:
                growth = self.a.revenue.growth_rates[year]
                revenue = revenue * (1 + growth)
                row["revenue"] = revenue

            # Revenue breakdown
            row["product_revenue"] = row["revenue"] * self.a.revenue.product_revenue_pct
            row["service_revenue"] = row["revenue"] * self.a.revenue.service_revenue_pct
            row["subscription_revenue"] = row["revenue"] * self.a.revenue.subscription_revenue_pct

            # Cost of Goods Sold
            row["cogs"] = row["revenue"] * self.a.cost.cogs_pct_of_revenue[year]
            row["gross_profit"] = row["revenue"] - row["cogs"]
            row["gross_margin"] = row["gross_profit"] / row["revenue"]

            # Operating Expenses
            row["sga"] = row["revenue"] * self.a.cost.sga_pct_of_revenue[year]
            row["rd"] = row["revenue"] * self.a.cost.rd_pct_of_revenue[year]
            row["depreciation"] = depreciation_schedule.get(year, 0.0)
            row["amortization"] = amortization_schedule.get(year, 0.0)
            row["total_opex"] = row["sga"] + row["rd"] + row["depreciation"] + row["amortization"]

            # Operating Income (EBIT)
            row["ebit"] = row["gross_profit"] - row["total_opex"]
            row["ebit_margin"] = row["ebit"] / row["revenue"]

            # Interest Expense
            row["interest_expense"] = interest_expense_schedule.get(year, 0.0)

            # Earnings Before Tax
            row["ebt"] = row["ebit"] - row["interest_expense"]

            # Tax
            row["tax_expense"] = max(row["ebt"] * self.a.balance_sheet.tax_rate, 0.0)

            # Net Income
            row["net_income"] = row["ebt"] - row["tax_expense"]
            row["net_margin"] = row["net_income"] / row["revenue"] if row["revenue"] != 0 else 0

            # EBITDA
            row["ebitda"] = row["ebit"] + row["depreciation"] + row["amortization"]
            row["ebitda_margin"] = row["ebitda"] / row["revenue"]

            self.data[year] = row

        return self.data


# ==============================================================================
# SECTION 3: BALANCE SHEET
# ==============================================================================

class BalanceSheet:
    """Builds the projected Balance Sheet."""

    def __init__(self, assumptions: Assumptions):
        self.a = assumptions
        self.data: Dict[int, Dict[str, float]] = {}

    def build_base_year(self) -> Dict[str, float]:
        bs = self.a.balance_sheet
        row = {}

        # Current Assets
        row["cash"] = bs.base_cash
        row["accounts_receivable"] = bs.base_accounts_receivable
        row["inventory"] = bs.base_inventory
        row["prepaid_expenses"] = bs.base_prepaid_expenses
        row["total_current_assets"] = (
            row["cash"] + row["accounts_receivable"]
            + row["inventory"] + row["prepaid_expenses"]
        )

        # Non-Current Assets
        row["ppe_gross"] = bs.base_ppe_gross
        row["accumulated_depreciation"] = bs.base_accumulated_depreciation
        row["ppe_net"] = row["ppe_gross"] - row["accumulated_depreciation"]
        row["intangibles_gross"] = bs.base_intangibles_gross
        row["accumulated_amortization"] = bs.base_accumulated_amortization
        row["intangibles_net"] = row["intangibles_gross"] - row["accumulated_amortization"]
        row["other_long_term_assets"] = bs.base_other_long_term_assets
        row["total_non_current_assets"] = (
            row["ppe_net"] + row["intangibles_net"] + row["other_long_term_assets"]
        )

        row["total_assets"] = row["total_current_assets"] + row["total_non_current_assets"]

        # Current Liabilities
        row["accounts_payable"] = bs.base_accounts_payable
        row["accrued_liabilities"] = bs.base_accrued_liabilities
        row["short_term_debt"] = bs.base_short_term_debt
        row["deferred_revenue"] = bs.base_deferred_revenue
        row["total_current_liabilities"] = (
            row["accounts_payable"] + row["accrued_liabilities"]
            + row["short_term_debt"] + row["deferred_revenue"]
        )

        # Non-Current Liabilities
        row["long_term_debt"] = bs.base_long_term_debt
        row["other_long_term_liabilities"] = bs.base_other_long_term_liabilities
        row["total_non_current_liabilities"] = (
            row["long_term_debt"] + row["other_long_term_liabilities"]
        )

        row["total_liabilities"] = (
            row["total_current_liabilities"] + row["total_non_current_liabilities"]
        )

        # Equity
        row["common_stock"] = bs.base_common_stock
        row["additional_paid_in_capital"] = bs.base_additional_paid_in_capital
        row["retained_earnings"] = bs.base_retained_earnings
        row["total_equity"] = (
            row["common_stock"] + row["additional_paid_in_capital"]
            + row["retained_earnings"]
        )

        row["total_liabilities_and_equity"] = row["total_liabilities"] + row["total_equity"]

        return row

    def build(self, income_data: Dict[int, Dict[str, float]],
              capex_schedule: Dict[int, float],
              intangible_investment_schedule: Dict[int, float],
              depreciation_schedule: Dict[int, float],
              amortization_schedule: Dict[int, float],
              dividends_schedule: Dict[int, float]) -> Dict[int, Dict[str, float]]:

        bs_a = self.a.balance_sheet

        # Base year
        self.data[BASE_YEAR] = self.build_base_year()

        for year in PROJECTED_YEARS:
            prev = self.data[year - 1]
            inc = income_data[year]
            row = {}

            revenue = inc["revenue"]
            cogs = inc["cogs"]
            total_opex = inc["total_opex"]

            # --- Current Assets ---
            row["accounts_receivable"] = revenue * (bs_a.days_sales_outstanding / 365)
            row["inventory"] = cogs * (bs_a.days_inventory_outstanding / 365)
            row["prepaid_expenses"] = revenue * bs_a.prepaid_pct_of_revenue

            # --- Non-Current Assets ---
            row["ppe_gross"] = prev["ppe_gross"] + capex_schedule[year]
            row["accumulated_depreciation"] = (
                prev["accumulated_depreciation"] + depreciation_schedule[year]
            )
            row["ppe_net"] = row["ppe_gross"] - row["accumulated_depreciation"]

            row["intangibles_gross"] = (
                prev["intangibles_gross"] + intangible_investment_schedule[year]
            )
            row["accumulated_amortization"] = (
                prev["accumulated_amortization"] + amortization_schedule[year]
            )
            row["intangibles_net"] = row["intangibles_gross"] - row["accumulated_amortization"]
            row["other_long_term_assets"] = prev["other_long_term_assets"]

            row["total_non_current_assets"] = (
                row["ppe_net"] + row["intangibles_net"] + row["other_long_term_assets"]
            )

            # --- Current Liabilities ---
            row["accounts_payable"] = cogs * (bs_a.days_payable_outstanding / 365)
            row["accrued_liabilities"] = total_opex * bs_a.accrued_liabilities_pct_of_opex
            row["short_term_debt"] = prev["short_term_debt"]  # Rolls forward
            row["deferred_revenue"] = prev["deferred_revenue"] * 1.05  # Slight growth
            row["total_current_liabilities"] = (
                row["accounts_payable"] + row["accrued_liabilities"]
                + row["short_term_debt"] + row["deferred_revenue"]
            )

            # --- Non-Current Liabilities ---
            new_debt = bs_a.new_debt_issuance.get(year, 0.0)
            row["long_term_debt"] = (
                prev["long_term_debt"] - bs_a.debt_repayment_per_year + new_debt
            )
            row["other_long_term_liabilities"] = prev["other_long_term_liabilities"]
            row["total_non_current_liabilities"] = (
                row["long_term_debt"] + row["other_long_term_liabilities"]
            )

            row["total_liabilities"] = (
                row["total_current_liabilities"] + row["total_non_current_liabilities"]
            )

            # --- Equity ---
            row["common_stock"] = prev["common_stock"]
            row["additional_paid_in_capital"] = prev["additional_paid_in_capital"]
            row["retained_earnings"] = (
                prev["retained_earnings"]
                + inc["net_income"]
                - dividends_schedule[year]
            )
            row["total_equity"] = (
                row["common_stock"] + row["additional_paid_in_capital"]
                + row["retained_earnings"]
            )

            row["total_liabilities_and_equity"] = row["total_liabilities"] + row["total_equity"]

            # --- Cash is the plug (balances the balance sheet) ---
            total_non_cash_current_assets = (
                row["accounts_receivable"] + row["inventory"] + row["prepaid_expenses"]
            )
            row["total_assets_excl_cash"] = (
                total_non_cash_current_assets + row["total_non_current_assets"]
            )
            row["cash"] = row["total_liabilities_and_equity"] - row["total_assets_excl_cash"]

            row["total_current_assets"] = (
                row["cash"] + row["accounts_receivable"]
                + row["inventory"] + row["prepaid_expenses"]
            )
            row["total_assets"] = row["total_current_assets"] + row["total_non_current_assets"]

            self.data[year] = row

        return self.data


# ==============================================================================
# SECTION 4: CASH FLOW STATEMENT
# ==============================================================================

class CashFlowStatement:
    """Builds the projected Cash Flow Statement."""

    def __init__(self, assumptions: Assumptions):
        self.a = assumptions
        self.data: Dict[int, Dict[str, float]] = {}

    def build(self, income_data: Dict[int, Dict[str, float]],
              bs_data: Dict[int, Dict[str, float]],
              capex_schedule: Dict[int, float],
              intangible_investment_schedule: Dict[int, float],
              depreciation_schedule: Dict[int, float],
              amortization_schedule: Dict[int, float],
              dividends_schedule: Dict[int, float]) -> Dict[int, Dict[str, float]]:

        bs_a = self.a.balance_sheet

        for year in PROJECTED_YEARS:
            prev_bs = bs_data[year - 1]
            curr_bs = bs_data[year]
            inc = income_data[year]
            row = {}

            # --- Operating Activities ---
            row["net_income"] = inc["net_income"]
            row["depreciation"] = depreciation_schedule[year]
            row["amortization"] = amortization_schedule[year]

            # Working capital changes (increase in asset = cash outflow)
            row["change_in_accounts_receivable"] = -(
                curr_bs["accounts_receivable"] - prev_bs["accounts_receivable"]
            )
            row["change_in_inventory"] = -(
                curr_bs["inventory"] - prev_bs["inventory"]
            )
            row["change_in_prepaid_expenses"] = -(
                curr_bs["prepaid_expenses"] - prev_bs["prepaid_expenses"]
            )
            row["change_in_accounts_payable"] = (
                curr_bs["accounts_payable"] - prev_bs["accounts_payable"]
            )
            row["change_in_accrued_liabilities"] = (
                curr_bs["accrued_liabilities"] - prev_bs["accrued_liabilities"]
            )
            row["change_in_deferred_revenue"] = (
                curr_bs["deferred_revenue"] - prev_bs["deferred_revenue"]
            )

            row["total_working_capital_changes"] = (
                row["change_in_accounts_receivable"]
                + row["change_in_inventory"]
                + row["change_in_prepaid_expenses"]
                + row["change_in_accounts_payable"]
                + row["change_in_accrued_liabilities"]
                + row["change_in_deferred_revenue"]
            )

            row["cash_from_operations"] = (
                row["net_income"]
                + row["depreciation"]
                + row["amortization"]
                + row["total_working_capital_changes"]
            )

            # --- Investing Activities ---
            row["capital_expenditures"] = -capex_schedule[year]
            row["intangible_investments"] = -intangible_investment_schedule[year]
            row["cash_from_investing"] = (
                row["capital_expenditures"] + row["intangible_investments"]
            )

            # --- Financing Activities ---
            new_debt = bs_a.new_debt_issuance.get(year, 0.0)
            row["debt_repayment"] = -bs_a.debt_repayment_per_year
            row["new_debt_issuance"] = new_debt
            row["dividends_paid"] = -dividends_schedule[year]
            row["cash_from_financing"] = (
                row["debt_repayment"] + row["new_debt_issuance"] + row["dividends_paid"]
            )

            # --- Net Change in Cash ---
            row["net_change_in_cash"] = (
                row["cash_from_operations"]
                + row["cash_from_investing"]
                + row["cash_from_financing"]
            )
            row["beginning_cash"] = prev_bs["cash"]
            row["ending_cash"] = row["beginning_cash"] + row["net_change_in_cash"]

            self.data[year] = row

        return self.data


# ==============================================================================
# SECTION 5: AUDIT MODULE
# ==============================================================================

class AuditResult:
    """Stores a single audit check result."""

    def __init__(self, check_name: str, passed: bool, year: Optional[int],
                 expected: float, actual: float, tolerance: float,
                 category: str, severity: str = "ERROR"):
        self.check_name = check_name
        self.passed = passed
        self.year = year
        self.expected = expected
        self.actual = actual
        self.difference = abs(expected - actual)
        self.tolerance = tolerance
        self.category = category
        self.severity = severity

    def __repr__(self):
        status = "PASS" if self.passed else f"**{self.severity}**"
        year_str = f" ({self.year})" if self.year else ""
        if self.passed:
            return f"  [{status}] {self.check_name}{year_str}"
        return (
            f"  [{status}] {self.check_name}{year_str}: "
            f"Expected={self.expected:,.2f}, Actual={self.actual:,.2f}, "
            f"Diff={self.difference:,.2f}"
        )


class Auditor:
    """
    Comprehensive audit of the 3-statement financial model.

    Audit Categories:
    1. Balance Sheet Balance Check (A = L + E)
    2. Cash Flow Reconciliation (ending cash matches BS cash)
    3. Retained Earnings Rollforward
    4. Working Capital Consistency
    5. Debt Schedule Reconciliation
    6. Cross-Statement Linkage Checks
    7. Ratio Reasonableness Checks
    8. Sign & Sanity Checks
    """

    TOLERANCE = 1.0  # $1 tolerance for rounding

    def __init__(self, income_data, bs_data, cf_data, assumptions,
                 depreciation_schedule, amortization_schedule,
                 dividends_schedule):
        self.income = income_data
        self.bs = bs_data
        self.cf = cf_data
        self.a = assumptions
        self.dep_sched = depreciation_schedule
        self.amort_sched = amortization_schedule
        self.div_sched = dividends_schedule
        self.results: List[AuditResult] = []
        self.summary = {}

    def _check(self, name, passed, year, expected, actual,
               category, severity="ERROR", tolerance=None):
        tol = tolerance if tolerance is not None else self.TOLERANCE
        self.results.append(AuditResult(
            name, passed, year, expected, actual, tol, category, severity
        ))

    def _approx_equal(self, a, b, tolerance=None):
        tol = tolerance if tolerance is not None else self.TOLERANCE
        return abs(a - b) <= tol

    def audit_all(self) -> List[AuditResult]:
        """Run all audit checks."""
        self.check_balance_sheet_balances()
        self.check_cash_flow_reconciliation()
        self.check_retained_earnings_rollforward()
        self.check_working_capital_consistency()
        self.check_debt_schedule()
        self.check_cross_statement_linkages()
        self.check_ratio_reasonableness()
        self.check_sign_sanity()
        self.check_growth_consistency()
        self.check_ppe_rollforward()
        self.check_intangibles_rollforward()

        # Build summary
        total = len(self.results)
        passed = sum(1 for r in self.results if r.passed)
        failed = total - passed
        errors = sum(1 for r in self.results if not r.passed and r.severity == "ERROR")
        warnings = sum(1 for r in self.results if not r.passed and r.severity == "WARNING")

        self.summary = {
            "total_checks": total,
            "passed": passed,
            "failed": failed,
            "errors": errors,
            "warnings": warnings,
            "pass_rate": f"{(passed / total * 100):.1f}%" if total > 0 else "N/A",
            "status": "CLEAN" if errors == 0 else "ISSUES FOUND",
        }

        return self.results

    # --- 1. Balance Sheet Balance ---
    def check_balance_sheet_balances(self):
        """Verify Assets = Liabilities + Equity for each year."""
        for year in YEARS:
            bs = self.bs[year]
            assets = bs["total_assets"]
            l_and_e = bs["total_liabilities_and_equity"]
            ok = self._approx_equal(assets, l_and_e)
            self._check(
                "Balance Sheet Balances (A = L + E)", ok, year,
                assets, l_and_e, "Balance Sheet"
            )

    # --- 2. Cash Flow Reconciliation ---
    def check_cash_flow_reconciliation(self):
        """Verify ending cash on CF matches BS cash."""
        for year in PROJECTED_YEARS:
            cf_ending = self.cf[year]["ending_cash"]
            bs_cash = self.bs[year]["cash"]
            ok = self._approx_equal(cf_ending, bs_cash)
            self._check(
                "CF Ending Cash = BS Cash", ok, year,
                bs_cash, cf_ending, "Cash Flow Reconciliation"
            )

    # --- 3. Retained Earnings Rollforward ---
    def check_retained_earnings_rollforward(self):
        """RE_end = RE_begin + Net Income - Dividends."""
        for year in PROJECTED_YEARS:
            prev_re = self.bs[year - 1]["retained_earnings"]
            ni = self.income[year]["net_income"]
            div = self.div_sched[year]
            expected_re = prev_re + ni - div
            actual_re = self.bs[year]["retained_earnings"]
            ok = self._approx_equal(expected_re, actual_re)
            self._check(
                "Retained Earnings Rollforward", ok, year,
                expected_re, actual_re, "Equity Reconciliation"
            )

    # --- 4. Working Capital Consistency ---
    def check_working_capital_consistency(self):
        """Verify working capital items are positive and reasonable."""
        for year in YEARS:
            bs = self.bs[year]
            for item in ["accounts_receivable", "inventory", "accounts_payable"]:
                val = bs[item]
                ok = val >= 0
                self._check(
                    f"{item} >= 0", ok, year, 0.0, val,
                    "Working Capital", "WARNING"
                )

    # --- 5. Debt Schedule ---
    def check_debt_schedule(self):
        """Verify debt rollforward: LTD_end = LTD_begin - repayment + new issuance."""
        for year in PROJECTED_YEARS:
            prev_ltd = self.bs[year - 1]["long_term_debt"]
            repayment = self.a.balance_sheet.debt_repayment_per_year
            new_debt = self.a.balance_sheet.new_debt_issuance.get(year, 0.0)
            expected = prev_ltd - repayment + new_debt
            actual = self.bs[year]["long_term_debt"]
            ok = self._approx_equal(expected, actual)
            self._check(
                "Long-Term Debt Rollforward", ok, year,
                expected, actual, "Debt Schedule"
            )

    # --- 6. Cross-Statement Linkages ---
    def check_cross_statement_linkages(self):
        """Verify IS-to-CF and IS-to-BS linkages."""
        for year in PROJECTED_YEARS:
            # Net income flows from IS to CF
            is_ni = self.income[year]["net_income"]
            cf_ni = self.cf[year]["net_income"]
            ok = self._approx_equal(is_ni, cf_ni)
            self._check(
                "IS Net Income = CF Net Income", ok, year,
                is_ni, cf_ni, "Cross-Statement Linkage"
            )

            # Depreciation flows from IS to CF
            is_dep = self.income[year]["depreciation"]
            cf_dep = self.cf[year]["depreciation"]
            ok = self._approx_equal(is_dep, cf_dep)
            self._check(
                "IS Depreciation = CF Depreciation", ok, year,
                is_dep, cf_dep, "Cross-Statement Linkage"
            )

            # Amortization flows from IS to CF
            is_amort = self.income[year]["amortization"]
            cf_amort = self.cf[year]["amortization"]
            ok = self._approx_equal(is_amort, cf_amort)
            self._check(
                "IS Amortization = CF Amortization", ok, year,
                is_amort, cf_amort, "Cross-Statement Linkage"
            )

    # --- 7. Ratio Reasonableness ---
    def check_ratio_reasonableness(self):
        """Check that key ratios are within reasonable bounds."""
        for year in YEARS:
            inc = self.income[year]

            # Gross margin between 40%-80%
            gm = inc["gross_margin"]
            ok = 0.20 <= gm <= 0.90
            self._check(
                "Gross Margin Reasonableness (20%-90%)", ok, year,
                0.55, gm, "Ratio Reasonableness", "WARNING", tolerance=0.35
            )

            # Net margin between -20% and 50%
            nm = inc["net_margin"]
            ok = -0.20 <= nm <= 0.50
            self._check(
                "Net Margin Reasonableness (-20% to 50%)", ok, year,
                0.15, nm, "Ratio Reasonableness", "WARNING", tolerance=0.35
            )

        # Debt-to-equity check
        for year in YEARS:
            bs = self.bs[year]
            if bs["total_equity"] > 0:
                de_ratio = bs["total_liabilities"] / bs["total_equity"]
                ok = de_ratio <= 5.0
                self._check(
                    "Debt-to-Equity <= 5.0x", ok, year,
                    5.0, de_ratio, "Ratio Reasonableness", "WARNING", tolerance=5.0
                )

    # --- 8. Sign & Sanity Checks ---
    def check_sign_sanity(self):
        """Verify that values have expected signs."""
        for year in YEARS:
            inc = self.income[year]
            bs = self.bs[year]

            # Revenue should be positive
            ok = inc["revenue"] > 0
            self._check("Revenue > 0", ok, year, 1.0, inc["revenue"],
                        "Sanity Check", "ERROR")

            # Total assets should be positive
            ok = bs["total_assets"] > 0
            self._check("Total Assets > 0", ok, year, 1.0, bs["total_assets"],
                        "Sanity Check", "ERROR")

            # Cash should be non-negative
            ok = bs["cash"] >= 0
            self._check("Cash >= 0", ok, year, 0.0, bs["cash"],
                        "Sanity Check", "WARNING")

            # Total equity should be positive (solvent company)
            ok = bs["total_equity"] > 0
            self._check("Total Equity > 0 (Solvency)", ok, year,
                        1.0, bs["total_equity"], "Sanity Check", "WARNING")

    # --- 9. Growth Consistency ---
    def check_growth_consistency(self):
        """Verify revenue growth matches assumptions."""
        for year in PROJECTED_YEARS:
            prev_rev = self.income[year - 1]["revenue"]
            curr_rev = self.income[year]["revenue"]
            actual_growth = (curr_rev - prev_rev) / prev_rev
            expected_growth = self.a.revenue.growth_rates[year]
            ok = self._approx_equal(actual_growth, expected_growth, tolerance=0.001)
            self._check(
                "Revenue Growth Matches Assumption", ok, year,
                expected_growth, actual_growth,
                "Growth Consistency", "ERROR", tolerance=0.001
            )

    # --- 10. PP&E Rollforward ---
    def check_ppe_rollforward(self):
        """Verify Gross PPE rollforward."""
        for year in PROJECTED_YEARS:
            prev_ppe = self.bs[year - 1]["ppe_gross"]
            capex = self.a.balance_sheet.capex_pct_of_revenue[year] * self.income[year]["revenue"]
            expected = prev_ppe + capex
            actual = self.bs[year]["ppe_gross"]
            ok = self._approx_equal(expected, actual)
            self._check(
                "Gross PP&E Rollforward", ok, year,
                expected, actual, "Fixed Asset Reconciliation"
            )

    # --- 11. Intangibles Rollforward ---
    def check_intangibles_rollforward(self):
        """Verify Gross Intangibles rollforward."""
        for year in PROJECTED_YEARS:
            prev = self.bs[year - 1]["intangibles_gross"]
            inv = (self.a.balance_sheet.intangible_investment_pct_of_revenue[year]
                   * self.income[year]["revenue"])
            expected = prev + inv
            actual = self.bs[year]["intangibles_gross"]
            ok = self._approx_equal(expected, actual)
            self._check(
                "Gross Intangibles Rollforward", ok, year,
                expected, actual, "Fixed Asset Reconciliation"
            )

    def print_report(self):
        """Print formatted audit report."""
        print("\n" + "=" * 80)
        print("  AUDIT REPORT - 5-Year Strategic Financial Model")
        print("=" * 80)

        # Group by category
        categories = {}
        for r in self.results:
            categories.setdefault(r.category, []).append(r)

        for cat, checks in categories.items():
            passed = sum(1 for c in checks if c.passed)
            total = len(checks)
            cat_status = "PASS" if passed == total else "ISSUES"
            print(f"\n  [{cat_status}] {cat} ({passed}/{total} passed)")
            print("  " + "-" * 60)
            for c in checks:
                print(f"  {c}")

        print("\n" + "=" * 80)
        print("  AUDIT SUMMARY")
        print("=" * 80)
        for k, v in self.summary.items():
            print(f"    {k:>20s}: {v}")
        print("=" * 80 + "\n")


# ==============================================================================
# SECTION 6: EXCEL OUTPUT
# ==============================================================================

def generate_excel(income_data, bs_data, cf_data, audit_results, audit_summary,
                   assumptions, filename="Strategic_Financial_Model_5Year.xlsx"):
    """Generate a professionally formatted Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # --- Style definitions ---
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    subheader_font = Font(name="Calibri", bold=True, size=10, color="2F5496")
    subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    total_font = Font(name="Calibri", bold=True, size=10)
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    number_format = '#,##0'
    pct_format = '0.0%'
    thin_border = Border(
        bottom=Side(style='thin', color='B4C6E7')
    )
    thick_border = Border(
        top=Side(style='medium', color='2F5496'),
        bottom=Side(style='medium', color='2F5496'),
    )
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    def write_header(ws, row, text, col_end):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        return row + 1

    def write_subheader(ws, row, text, col_end):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = subheader_font
        cell.fill = subheader_fill
        return row + 1

    def write_year_headers(ws, row, years):
        ws.cell(row=row, column=1, value="").font = Font(bold=True)
        for i, y in enumerate(years):
            label = f"FY{y}" if y != BASE_YEAR else f"FY{y} (Base)"
            cell = ws.cell(row=row, column=i + 2, value=label)
            cell.font = Font(name="Calibri", bold=True, size=10, color="2F5496")
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(bottom=Side(style='medium', color='2F5496'))
        return row + 1

    def write_line(ws, row, label, values, fmt=number_format, is_total=False, years=YEARS):
        cell = ws.cell(row=row, column=1, value=label)
        if is_total:
            cell.font = total_font
        for i, y in enumerate(years):
            c = ws.cell(row=row, column=i + 2, value=values.get(y, 0))
            c.number_format = fmt
            if is_total:
                c.font = total_font
                c.fill = total_fill
                c.border = thick_border
            else:
                c.border = thin_border
        return row + 1

    col_end = len(YEARS) + 1

    # ========== ASSUMPTIONS SHEET ==========
    ws_a = wb.active
    ws_a.title = "Assumptions"
    ws_a.column_dimensions['A'].width = 40
    for i in range(2, col_end + 1):
        ws_a.column_dimensions[get_column_letter(i)].width = 16

    r = 1
    r = write_header(ws_a, r, "ASSUMPTIONS & DRIVERS", col_end)
    r += 1

    r = write_subheader(ws_a, r, "Revenue Assumptions", col_end)
    r = write_year_headers(ws_a, r, YEARS)
    rev = assumptions.revenue.base_revenue
    rev_vals = {BASE_YEAR: rev}
    for y in PROJECTED_YEARS:
        rev = rev * (1 + assumptions.revenue.growth_rates[y])
        rev_vals[y] = rev
    r = write_line(ws_a, r, "Revenue", rev_vals)
    r = write_line(ws_a, r, "Revenue Growth Rate",
                   {BASE_YEAR: 0, **assumptions.revenue.growth_rates}, fmt=pct_format)
    r += 1

    r = write_subheader(ws_a, r, "Cost Assumptions (% of Revenue)", col_end)
    r = write_line(ws_a, r, "COGS %", assumptions.cost.cogs_pct_of_revenue, fmt=pct_format)
    r = write_line(ws_a, r, "SG&A %", assumptions.cost.sga_pct_of_revenue, fmt=pct_format)
    r = write_line(ws_a, r, "R&D %", assumptions.cost.rd_pct_of_revenue, fmt=pct_format)
    r += 1

    r = write_subheader(ws_a, r, "Balance Sheet Assumptions", col_end)
    ws_a.cell(row=r, column=1, value="Days Sales Outstanding (DSO)")
    ws_a.cell(row=r, column=2, value=assumptions.balance_sheet.days_sales_outstanding)
    r += 1
    ws_a.cell(row=r, column=1, value="Days Inventory Outstanding (DIO)")
    ws_a.cell(row=r, column=2, value=assumptions.balance_sheet.days_inventory_outstanding)
    r += 1
    ws_a.cell(row=r, column=1, value="Days Payable Outstanding (DPO)")
    ws_a.cell(row=r, column=2, value=assumptions.balance_sheet.days_payable_outstanding)
    r += 1
    ws_a.cell(row=r, column=1, value="Tax Rate")
    c = ws_a.cell(row=r, column=2, value=assumptions.balance_sheet.tax_rate)
    c.number_format = pct_format
    r += 1
    ws_a.cell(row=r, column=1, value="Interest Rate")
    c = ws_a.cell(row=r, column=2, value=assumptions.balance_sheet.interest_rate)
    c.number_format = pct_format
    r += 1
    ws_a.cell(row=r, column=1, value="Dividend Payout Ratio")
    c = ws_a.cell(row=r, column=2, value=assumptions.balance_sheet.dividend_payout_ratio)
    c.number_format = pct_format
    r += 1
    r = write_line(ws_a, r, "CapEx (% of Revenue)",
                   {BASE_YEAR: 0, **assumptions.balance_sheet.capex_pct_of_revenue}, fmt=pct_format)

    # ========== INCOME STATEMENT SHEET ==========
    ws_is = wb.create_sheet("Income Statement")
    ws_is.column_dimensions['A'].width = 40
    for i in range(2, col_end + 1):
        ws_is.column_dimensions[get_column_letter(i)].width = 16

    r = 1
    r = write_header(ws_is, r, "INCOME STATEMENT", col_end)
    r = write_year_headers(ws_is, r, YEARS)

    def get_is_vals(key):
        return {y: income_data[y][key] for y in YEARS}

    r = write_subheader(ws_is, r, "Revenue", col_end)
    r = write_line(ws_is, r, "  Product Revenue", get_is_vals("product_revenue"))
    r = write_line(ws_is, r, "  Service Revenue", get_is_vals("service_revenue"))
    r = write_line(ws_is, r, "  Subscription Revenue", get_is_vals("subscription_revenue"))
    r = write_line(ws_is, r, "Total Revenue", get_is_vals("revenue"), is_total=True)
    r += 1

    r = write_subheader(ws_is, r, "Cost of Goods Sold", col_end)
    r = write_line(ws_is, r, "  COGS", get_is_vals("cogs"))
    r = write_line(ws_is, r, "Gross Profit", get_is_vals("gross_profit"), is_total=True)
    r = write_line(ws_is, r, "  Gross Margin %", get_is_vals("gross_margin"), fmt=pct_format)
    r += 1

    r = write_subheader(ws_is, r, "Operating Expenses", col_end)
    r = write_line(ws_is, r, "  Selling, General & Administrative", get_is_vals("sga"))
    r = write_line(ws_is, r, "  Research & Development", get_is_vals("rd"))
    r = write_line(ws_is, r, "  Depreciation", get_is_vals("depreciation"))
    r = write_line(ws_is, r, "  Amortization", get_is_vals("amortization"))
    r = write_line(ws_is, r, "Total Operating Expenses", get_is_vals("total_opex"), is_total=True)
    r += 1

    r = write_line(ws_is, r, "EBIT (Operating Income)", get_is_vals("ebit"), is_total=True)
    r = write_line(ws_is, r, "  EBIT Margin %", get_is_vals("ebit_margin"), fmt=pct_format)
    r = write_line(ws_is, r, "  Interest Expense", get_is_vals("interest_expense"))
    r = write_line(ws_is, r, "Earnings Before Tax", get_is_vals("ebt"), is_total=True)
    r = write_line(ws_is, r, "  Tax Expense (25%)", get_is_vals("tax_expense"))
    r += 1
    r = write_line(ws_is, r, "NET INCOME", get_is_vals("net_income"), is_total=True)
    r = write_line(ws_is, r, "  Net Margin %", get_is_vals("net_margin"), fmt=pct_format)
    r += 1
    r = write_line(ws_is, r, "EBITDA", get_is_vals("ebitda"), is_total=True)
    r = write_line(ws_is, r, "  EBITDA Margin %", get_is_vals("ebitda_margin"), fmt=pct_format)

    # ========== BALANCE SHEET ==========
    ws_bs = wb.create_sheet("Balance Sheet")
    ws_bs.column_dimensions['A'].width = 40
    for i in range(2, col_end + 1):
        ws_bs.column_dimensions[get_column_letter(i)].width = 16

    r = 1
    r = write_header(ws_bs, r, "BALANCE SHEET", col_end)
    r = write_year_headers(ws_bs, r, YEARS)

    def get_bs_vals(key):
        return {y: bs_data[y][key] for y in YEARS}

    r = write_subheader(ws_bs, r, "ASSETS", col_end)
    r = write_subheader(ws_bs, r, "Current Assets", col_end)
    r = write_line(ws_bs, r, "  Cash & Cash Equivalents", get_bs_vals("cash"))
    r = write_line(ws_bs, r, "  Accounts Receivable", get_bs_vals("accounts_receivable"))
    r = write_line(ws_bs, r, "  Inventory", get_bs_vals("inventory"))
    r = write_line(ws_bs, r, "  Prepaid Expenses", get_bs_vals("prepaid_expenses"))
    r = write_line(ws_bs, r, "Total Current Assets", get_bs_vals("total_current_assets"), is_total=True)
    r += 1

    r = write_subheader(ws_bs, r, "Non-Current Assets", col_end)
    r = write_line(ws_bs, r, "  PP&E (Gross)", get_bs_vals("ppe_gross"))
    r = write_line(ws_bs, r, "  Less: Accumulated Depreciation", get_bs_vals("accumulated_depreciation"))
    r = write_line(ws_bs, r, "  PP&E (Net)", get_bs_vals("ppe_net"))
    r = write_line(ws_bs, r, "  Intangibles (Gross)", get_bs_vals("intangibles_gross"))
    r = write_line(ws_bs, r, "  Less: Accumulated Amortization", get_bs_vals("accumulated_amortization"))
    r = write_line(ws_bs, r, "  Intangibles (Net)", get_bs_vals("intangibles_net"))
    r = write_line(ws_bs, r, "  Other Long-Term Assets", get_bs_vals("other_long_term_assets"))
    r = write_line(ws_bs, r, "Total Non-Current Assets", get_bs_vals("total_non_current_assets"), is_total=True)
    r += 1
    r = write_line(ws_bs, r, "TOTAL ASSETS", get_bs_vals("total_assets"), is_total=True)
    r += 1

    r = write_subheader(ws_bs, r, "LIABILITIES", col_end)
    r = write_subheader(ws_bs, r, "Current Liabilities", col_end)
    r = write_line(ws_bs, r, "  Accounts Payable", get_bs_vals("accounts_payable"))
    r = write_line(ws_bs, r, "  Accrued Liabilities", get_bs_vals("accrued_liabilities"))
    r = write_line(ws_bs, r, "  Short-Term Debt", get_bs_vals("short_term_debt"))
    r = write_line(ws_bs, r, "  Deferred Revenue", get_bs_vals("deferred_revenue"))
    r = write_line(ws_bs, r, "Total Current Liabilities", get_bs_vals("total_current_liabilities"), is_total=True)
    r += 1

    r = write_subheader(ws_bs, r, "Non-Current Liabilities", col_end)
    r = write_line(ws_bs, r, "  Long-Term Debt", get_bs_vals("long_term_debt"))
    r = write_line(ws_bs, r, "  Other Long-Term Liabilities", get_bs_vals("other_long_term_liabilities"))
    r = write_line(ws_bs, r, "Total Non-Current Liabilities", get_bs_vals("total_non_current_liabilities"), is_total=True)
    r += 1
    r = write_line(ws_bs, r, "TOTAL LIABILITIES", get_bs_vals("total_liabilities"), is_total=True)
    r += 1

    r = write_subheader(ws_bs, r, "SHAREHOLDERS' EQUITY", col_end)
    r = write_line(ws_bs, r, "  Common Stock", get_bs_vals("common_stock"))
    r = write_line(ws_bs, r, "  Additional Paid-In Capital", get_bs_vals("additional_paid_in_capital"))
    r = write_line(ws_bs, r, "  Retained Earnings", get_bs_vals("retained_earnings"))
    r = write_line(ws_bs, r, "Total Shareholders' Equity", get_bs_vals("total_equity"), is_total=True)
    r += 1
    r = write_line(ws_bs, r, "TOTAL LIABILITIES & EQUITY", get_bs_vals("total_liabilities_and_equity"), is_total=True)

    # Balance check row
    r += 1
    ws_bs.cell(row=r, column=1, value="Balance Check (Assets - L&E)").font = Font(bold=True, color="FF0000")
    for i, y in enumerate(YEARS):
        diff = bs_data[y]["total_assets"] - bs_data[y]["total_liabilities_and_equity"]
        c = ws_bs.cell(row=r, column=i + 2, value=diff)
        c.number_format = '#,##0.00'
        c.font = Font(bold=True, color="006100" if abs(diff) < 1 else "FF0000")
        c.fill = pass_fill if abs(diff) < 1 else fail_fill

    # ========== CASH FLOW STATEMENT ==========
    ws_cf = wb.create_sheet("Cash Flow Statement")
    ws_cf.column_dimensions['A'].width = 40
    for i in range(2, len(PROJECTED_YEARS) + 2):
        ws_cf.column_dimensions[get_column_letter(i)].width = 16

    cf_col_end = len(PROJECTED_YEARS) + 1
    r = 1
    r = write_header(ws_cf, r, "CASH FLOW STATEMENT", cf_col_end)
    r = write_year_headers(ws_cf, r, PROJECTED_YEARS)

    def get_cf_vals(key):
        return {y: cf_data[y][key] for y in PROJECTED_YEARS}

    r = write_subheader(ws_cf, r, "Cash from Operating Activities", cf_col_end)
    r = write_line(ws_cf, r, "  Net Income", get_cf_vals("net_income"), years=PROJECTED_YEARS)
    r = write_line(ws_cf, r, "  Depreciation", get_cf_vals("depreciation"), years=PROJECTED_YEARS)
    r = write_line(ws_cf, r, "  Amortization", get_cf_vals("amortization"), years=PROJECTED_YEARS)
    r = write_line(ws_cf, r, "  Change in Accounts Receivable", get_cf_vals("change_in_accounts_receivable"), years=PROJECTED_YEARS)
    r = write_line(ws_cf, r, "  Change in Inventory", get_cf_vals("change_in_inventory"), years=PROJECTED_YEARS)
    r = write_line(ws_cf, r, "  Change in Prepaid Expenses", get_cf_vals("change_in_prepaid_expenses"), years=PROJECTED_YEARS)
    r = write_line(ws_cf, r, "  Change in Accounts Payable", get_cf_vals("change_in_accounts_payable"), years=PROJECTED_YEARS)
    r = write_line(ws_cf, r, "  Change in Accrued Liabilities", get_cf_vals("change_in_accrued_liabilities"), years=PROJECTED_YEARS)
    r = write_line(ws_cf, r, "  Change in Deferred Revenue", get_cf_vals("change_in_deferred_revenue"), years=PROJECTED_YEARS)
    r = write_line(ws_cf, r, "Cash from Operations", get_cf_vals("cash_from_operations"), is_total=True, years=PROJECTED_YEARS)
    r += 1

    r = write_subheader(ws_cf, r, "Cash from Investing Activities", cf_col_end)
    r = write_line(ws_cf, r, "  Capital Expenditures", get_cf_vals("capital_expenditures"), years=PROJECTED_YEARS)
    r = write_line(ws_cf, r, "  Intangible Investments", get_cf_vals("intangible_investments"), years=PROJECTED_YEARS)
    r = write_line(ws_cf, r, "Cash from Investing", get_cf_vals("cash_from_investing"), is_total=True, years=PROJECTED_YEARS)
    r += 1

    r = write_subheader(ws_cf, r, "Cash from Financing Activities", cf_col_end)
    r = write_line(ws_cf, r, "  Debt Repayment", get_cf_vals("debt_repayment"), years=PROJECTED_YEARS)
    r = write_line(ws_cf, r, "  New Debt Issuance", get_cf_vals("new_debt_issuance"), years=PROJECTED_YEARS)
    r = write_line(ws_cf, r, "  Dividends Paid", get_cf_vals("dividends_paid"), years=PROJECTED_YEARS)
    r = write_line(ws_cf, r, "Cash from Financing", get_cf_vals("cash_from_financing"), is_total=True, years=PROJECTED_YEARS)
    r += 1

    r = write_line(ws_cf, r, "NET CHANGE IN CASH", get_cf_vals("net_change_in_cash"), is_total=True, years=PROJECTED_YEARS)
    r = write_line(ws_cf, r, "Beginning Cash", get_cf_vals("beginning_cash"), years=PROJECTED_YEARS)
    r = write_line(ws_cf, r, "ENDING CASH", get_cf_vals("ending_cash"), is_total=True, years=PROJECTED_YEARS)

    # ========== AUDIT SHEET ==========
    ws_audit = wb.create_sheet("Audit Report")
    ws_audit.column_dimensions['A'].width = 10
    ws_audit.column_dimensions['B'].width = 40
    ws_audit.column_dimensions['C'].width = 10
    ws_audit.column_dimensions['D'].width = 12
    ws_audit.column_dimensions['E'].width = 18
    ws_audit.column_dimensions['F'].width = 18
    ws_audit.column_dimensions['G'].width = 18
    ws_audit.column_dimensions['H'].width = 14

    r = 1
    ws_audit.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    c = ws_audit.cell(row=r, column=1, value="AUDIT REPORT - 5-Year Strategic Financial Model")
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center')
    r += 1

    # Summary
    r = write_subheader(ws_audit, r, "Audit Summary", 8)
    for k, v in audit_summary.items():
        ws_audit.cell(row=r, column=2, value=k.replace("_", " ").title())
        ws_audit.cell(row=r, column=3, value=str(v))
        if k == "status":
            ws_audit.cell(row=r, column=3).fill = (
                pass_fill if v == "CLEAN" else fail_fill
            )
        r += 1
    r += 1

    # Detail
    r = write_subheader(ws_audit, r, "Detailed Audit Results", 8)
    headers = ["Status", "Check Name", "Year", "Category", "Expected", "Actual", "Difference", "Severity"]
    for i, h in enumerate(headers):
        c = ws_audit.cell(row=r, column=i + 1, value=h)
        c.font = Font(bold=True)
        c.fill = subheader_fill
    r += 1

    for ar in audit_results:
        status = "PASS" if ar.passed else "FAIL"
        ws_audit.cell(row=r, column=1, value=status)
        ws_audit.cell(row=r, column=1).fill = pass_fill if ar.passed else (
            fail_fill if ar.severity == "ERROR" else warn_fill
        )
        ws_audit.cell(row=r, column=2, value=ar.check_name)
        ws_audit.cell(row=r, column=3, value=ar.year if ar.year else "")
        ws_audit.cell(row=r, column=4, value=ar.category)
        ws_audit.cell(row=r, column=5, value=ar.expected).number_format = '#,##0.00'
        ws_audit.cell(row=r, column=6, value=ar.actual).number_format = '#,##0.00'
        ws_audit.cell(row=r, column=7, value=ar.difference).number_format = '#,##0.00'
        ws_audit.cell(row=r, column=8, value=ar.severity)
        r += 1

    wb.save(filename)
    return filename


# ==============================================================================
# SECTION 7: MODEL ORCHESTRATOR
# ==============================================================================

class FinancialModel:
    """Orchestrates the 3-statement model build and audit."""

    def __init__(self, assumptions: Optional[Assumptions] = None):
        self.a = assumptions or Assumptions()
        self.income_data = {}
        self.bs_data = {}
        self.cf_data = {}
        self.audit_results = []
        self.audit_summary = {}

    def _build_schedules(self):
        """Pre-compute depreciation, amortization, capex, interest, dividends."""
        # We need a first pass of revenue to compute capex, etc.
        revenue = self.a.revenue.base_revenue
        revenue_by_year = {BASE_YEAR: revenue}
        for y in PROJECTED_YEARS:
            revenue = revenue * (1 + self.a.revenue.growth_rates[y])
            revenue_by_year[y] = revenue

        # CapEx schedule
        self.capex_schedule = {}
        for y in PROJECTED_YEARS:
            self.capex_schedule[y] = (
                self.a.balance_sheet.capex_pct_of_revenue[y] * revenue_by_year[y]
            )

        # Intangible investment schedule
        self.intangible_schedule = {}
        for y in PROJECTED_YEARS:
            self.intangible_schedule[y] = (
                self.a.balance_sheet.intangible_investment_pct_of_revenue[y]
                * revenue_by_year[y]
            )

        # PP&E Gross rollforward for depreciation calc
        ppe_gross = self.a.balance_sheet.base_ppe_gross
        self.depreciation_schedule = {
            BASE_YEAR: ppe_gross * self.a.cost.depreciation_pct_of_ppe
        }
        for y in PROJECTED_YEARS:
            ppe_gross += self.capex_schedule[y]
            self.depreciation_schedule[y] = ppe_gross * self.a.cost.depreciation_pct_of_ppe

        # Intangibles Gross rollforward for amortization calc
        intangibles_gross = self.a.balance_sheet.base_intangibles_gross
        self.amortization_schedule = {
            BASE_YEAR: intangibles_gross * self.a.cost.amortization_pct_of_intangibles
        }
        for y in PROJECTED_YEARS:
            intangibles_gross += self.intangible_schedule[y]
            self.amortization_schedule[y] = (
                intangibles_gross * self.a.cost.amortization_pct_of_intangibles
            )

        # Interest expense (based on average debt)
        ltd = self.a.balance_sheet.base_long_term_debt
        std = self.a.balance_sheet.base_short_term_debt
        self.interest_schedule = {
            BASE_YEAR: (ltd + std) * self.a.balance_sheet.interest_rate
        }
        for y in PROJECTED_YEARS:
            new_debt = self.a.balance_sheet.new_debt_issuance.get(y, 0.0)
            end_ltd = ltd - self.a.balance_sheet.debt_repayment_per_year + new_debt
            avg_debt = ((ltd + end_ltd) / 2) + std
            self.interest_schedule[y] = avg_debt * self.a.balance_sheet.interest_rate
            ltd = end_ltd

        # Dividends (computed after income statement - placeholder)
        self.dividends_schedule = {}

    def build(self):
        """Build the full 3-statement model."""
        print("Building 5-Year Strategic Financial Model...")
        print("-" * 50)

        # Step 1: Pre-compute schedules
        print("  [1/6] Computing assumptions & schedules...")
        self._build_schedules()

        # Step 2: Build Income Statement
        print("  [2/6] Building Income Statement...")
        is_builder = IncomeStatement(self.a)
        self.income_data = is_builder.build(
            self.depreciation_schedule,
            self.amortization_schedule,
            self.interest_schedule,
        )

        # Step 3: Compute dividends from net income
        print("  [3/6] Computing dividend schedule...")
        for y in PROJECTED_YEARS:
            ni = self.income_data[y]["net_income"]
            self.dividends_schedule[y] = max(
                ni * self.a.balance_sheet.dividend_payout_ratio, 0.0
            )

        # Step 4: Build Balance Sheet
        print("  [4/6] Building Balance Sheet...")
        bs_builder = BalanceSheet(self.a)
        self.bs_data = bs_builder.build(
            self.income_data,
            self.capex_schedule,
            self.intangible_schedule,
            self.depreciation_schedule,
            self.amortization_schedule,
            self.dividends_schedule,
        )

        # Step 5: Build Cash Flow Statement
        print("  [5/6] Building Cash Flow Statement...")
        cf_builder = CashFlowStatement(self.a)
        self.cf_data = cf_builder.build(
            self.income_data,
            self.bs_data,
            self.capex_schedule,
            self.intangible_schedule,
            self.depreciation_schedule,
            self.amortization_schedule,
            self.dividends_schedule,
        )

        # Step 6: Audit
        print("  [6/6] Running audit...")
        auditor = Auditor(
            self.income_data, self.bs_data, self.cf_data, self.a,
            self.depreciation_schedule, self.amortization_schedule,
            self.dividends_schedule,
        )
        self.audit_results = auditor.audit_all()
        self.audit_summary = auditor.summary
        auditor.print_report()

        # Generate Excel
        print("\nGenerating Excel workbook...")
        filename = generate_excel(
            self.income_data, self.bs_data, self.cf_data,
            self.audit_results, self.audit_summary, self.a,
        )
        print(f"  -> Saved to: {filename}")

        # Print key metrics summary
        self._print_key_metrics()

        return filename

    def _print_key_metrics(self):
        """Print a summary of key strategic metrics."""
        print("\n" + "=" * 80)
        print("  KEY STRATEGIC METRICS SUMMARY")
        print("=" * 80)
        print(f"\n  {'Metric':<35s}", end="")
        for y in YEARS:
            print(f"{'FY' + str(y):>12s}", end="")
        print()
        print("  " + "-" * (35 + 12 * len(YEARS)))

        metrics = [
            ("Revenue ($)", "revenue", "dollar"),
            ("Revenue Growth (%)", None, "growth"),
            ("Gross Margin (%)", "gross_margin", "pct"),
            ("EBITDA ($)", "ebitda", "dollar"),
            ("EBITDA Margin (%)", "ebitda_margin", "pct"),
            ("Net Income ($)", "net_income", "dollar"),
            ("Net Margin (%)", "net_margin", "pct"),
        ]

        for label, key, fmt in metrics:
            print(f"  {label:<35s}", end="")
            for y in YEARS:
                if fmt == "growth":
                    if y == BASE_YEAR:
                        print(f"{'N/A':>12s}", end="")
                    else:
                        g = self.a.revenue.growth_rates[y]
                        print(f"{g:>11.1%} ", end="")
                elif fmt == "dollar":
                    val = self.income_data[y][key]
                    print(f"{val:>12,.0f}", end="")
                elif fmt == "pct":
                    val = self.income_data[y][key]
                    print(f"{val:>11.1%} ", end="")
            print()

        # Balance sheet highlights
        print()
        bs_metrics = [
            ("Total Assets ($)", "total_assets"),
            ("Total Debt ($)", None),
            ("Total Equity ($)", "total_equity"),
            ("Cash ($)", "cash"),
        ]

        for label, key in bs_metrics:
            print(f"  {label:<35s}", end="")
            for y in YEARS:
                if key:
                    val = self.bs_data[y][key]
                elif "Debt" in label:
                    val = self.bs_data[y]["long_term_debt"] + self.bs_data[y]["short_term_debt"]
                else:
                    val = 0
                print(f"{val:>12,.0f}", end="")
            print()

        # Cash flow highlights
        print()
        print(f"  {'Cash from Operations ($)':<35s}", end="")
        for y in PROJECTED_YEARS:
            val = self.cf_data[y]["cash_from_operations"]
            print(f"{val:>12,.0f}", end="")
        print()
        print(f"  {'Free Cash Flow ($)':<35s}", end="")
        for y in PROJECTED_YEARS:
            cfo = self.cf_data[y]["cash_from_operations"]
            capex = self.capex_schedule[y]
            fcf = cfo - capex
            print(f"{fcf:>12,.0f}", end="")
        print()

        print("\n" + "=" * 80)


# ==============================================================================
# MAIN EXECUTION
# ==============================================================================

if __name__ == "__main__":
    model = FinancialModel()
    output_file = model.build()
    print(f"\nModel complete. Output: {output_file}")
