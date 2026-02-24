import openpyxl
import re
import json
from collections import Counter, defaultdict
from spellchecker import SpellChecker
from openpyxl.styles import PatternFill
from openpyxl.worksheet.formula import ArrayFormula

def get_color_str(color):
    if color is None: return "None"
    if color.type == 'rgb': return f"#{color.rgb[2:]}"
    elif color.type == 'theme': return f"Theme({color.theme})"
    elif color.type == 'indexed': return f"Indexed({color.indexed})"
    return "None"

def a1_to_r1c1(formula, cell_row, cell_col):
    if isinstance(formula, ArrayFormula): formula = formula.t
    formula = str(formula)
    def replace_ref(match):
        col_str, row_str = match.groups()
        col_abs = col_str.startswith('$')
        row_abs = row_str.startswith('$')
        col_name = col_str.replace('$', '')
        row_num = int(row_str.replace('$', ''))
        col_index = 0
        for char in col_name: col_index = col_index * 26 + (ord(char.upper()) - ord('A') + 1)
        if row_abs: r_part = f"R{row_num}"
        else:
            r_part = f"R[{row_num - cell_row}]"
            if r_part == "R[0]": r_part = "R"
        if col_abs: c_part = f"C{col_index}"
        else:
            c_part = f"C[{col_index - cell_col}]"
            if c_part == "C[0]": c_part = "C"
        return r_part + c_part
    pattern = r'(\$?[A-Z]+)(\$?[0-9]+)'
    return re.sub(pattern, replace_ref, formula)

def full_audit(filename):
    wb = openpyxl.load_workbook(filename, data_only=False)
    wb_data = openpyxl.load_workbook(filename, data_only=True)
    findings = []

    # 1. Sentry & Efficiency & Architect
    error_values = ["#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#SPILL!", "#CALC!"]
    volatile_functions = ["OFFSET(", "INDIRECT(", "TODAY()", "NOW(", "RAND(", "RANDBETWEEN("]

    # Identify Timeline Sheet for Architect
    timeline_sheet = next((s for s in wb.sheetnames if 'timeline' in s.lower() or 'timing' in s.lower()), None)

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        ws_data = wb_data[sheetname]
        for row in ws.iter_rows():
            for cell in row:
                val = ws_data[cell.coordinate].value
                if str(val) in error_values:
                    findings.append({"Sheet Name": sheetname, "Cell Reference": cell.coordinate, "Description": "Calculation Error", "Category": "Calculation Error", "Long Description": f"🔴 HIGH: Cell contains error {val}"})

                if cell.data_type == 'f':
                    formula = str(cell.value).upper()

                    # Efficiency: Volatile functions
                    for vf in volatile_functions:
                        if vf in formula:
                            findings.append({
                                "Sheet Name": sheetname, "Cell Reference": cell.coordinate,
                                "Description": f"Volatile Function: {vf[:-1]}",
                                "Category": "Volatile Complexity",
                                "Long Description": f"⚠️ MEDIUM: Use of volatile function {vf[:-1]} detected. This can cause performance degradation in large models."
                            })

                    # Efficiency: Mega-Formula (>4000)
                    if len(formula) > 4000:
                        findings.append({
                            "Sheet Name": sheetname, "Cell Reference": cell.coordinate,
                            "Description": "Mega-Formula (Bad Practice)",
                            "Category": "Mega-Formula",
                            "Long Description": f"🔴 HIGH: Extremely long formula detected (>4000 chars). Flagged as CRITICAL BAD PRACTICE due to total lack of auditability."
                        })
                    # Efficiency: Long Formula (>500)
                    elif len(formula) > 500:
                        findings.append({
                            "Sheet Name": sheetname, "Cell Reference": cell.coordinate,
                            "Description": "Long Formula",
                            "Category": "Auditability Risk",
                            "Long Description": f"⚠️ MEDIUM: Formula exceeds 500 characters. While not a 'Mega-Formula', it still poses an auditability risk and should be broken down."
                        })

                    # Architect: Master Date Spine alignment
                    if timeline_sheet and sheetname != timeline_sheet:
                        if '!' in formula and timeline_sheet not in formula:
                            # If it links to other sheets but NOT the timeline sheet, it might be an architectural risk
                            # but this is too noisy. Let's look for hardcoded dates or inconsistent period references.
                            pass

                    # Architect: Scalability (Check for hardcoded column limits in SUM/etc)
                    if any(re.search(r'[A-Z]+\$[0-9]+:[A-Z]+\$[0-9]+', formula) for _ in [1]):
                         # This is complex to automate perfectly, but we can flag some patterns
                         pass

    # 2. Lingo
    spell = SpellChecker()
    spell.word_frequency.load_words(['EBITDA', 'COGS', 'CFADS', 'IRR', 'CAPEX', 'OPEX', 'GST', 'Corality', 'Excel', 'tutorial', 'modelling', 'Offsheet', 'OneLine', 'DSCR', 'Heng'])
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        for row in ws.iter_rows(max_row=30):
            for cell in row:
                if isinstance(cell.value, str) and not cell.value.startswith('=') and cell.column <= 5:
                    words = re.findall(r'\b[A-Za-z]{4,}\b', cell.value)
                    misspelled = spell.unknown(words)
                    for word in misspelled:
                        if word.lower() not in ['dscr', 'heng']:
                            findings.append({"Sheet Name": sheetname, "Cell Reference": cell.coordinate, "Description": f"Label: {cell.value}", "Category": "Typo", "Long Description": f"🟡 LOW: Potential typo '{word}' in '{cell.value}'"})

    # 3. Stylist
    stylist_findings = defaultdict(list)
    for sheetname in wb.sheetnames:
        if sheetname in ['L', '  ', 'Corality']: continue
        ws = wb[sheetname]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None: continue
                f_color = get_color_str(cell.font.color)
                if not cell.data_type == 'f' and isinstance(cell.value, (int, float)):
                    if cell.row > 5 and cell.column > 3:
                        if f_color != "#800000" and f_color != "None" and f_color != "Theme(0)":
                            stylist_findings[(sheetname, "Colour Coding Error", f"🟡 LOW: Hard-coded input is formatted with font color {f_color} instead of the model's Input style (#800000)")].append(cell.coordinate)
                if cell.data_type == 'f' and '!' in str(cell.value):
                    if f_color != "Indexed(16)":
                        stylist_findings[(sheetname, "Colour Coding Error", "🟡 LOW: Off-sheet link is formatted with font color " + f_color + " instead of the model's Link style (Indexed 16)")].append(cell.coordinate)
    for (sn, cat, desc), cells in stylist_findings.items():
        findings.append({"Sheet Name": sn, "Cell Reference": ", ".join(cells), "Description": "Multiple locations", "Category": cat, "Long Description": desc})

    # 4. Logic & Numerical Sense Check
    for sheetname in wb.sheetnames:
        if sheetname in ['L', '  ', 'Corality']: continue
        ws = wb[sheetname]
        ws_data = wb_data[sheetname]

        row_labels = {}
        for row in ws.iter_rows(min_col=1, max_col=5):
            label = ""
            for cell in row:
                if isinstance(cell.value, str): label = cell.value; break
            row_labels[row[0].row] = label

        for row_idx, label in row_labels.items():
            label_lower = label.lower()
            for cell in ws_data[row_idx]:
                val = cell.value
                if isinstance(val, (int, float)):
                    if "dscr" in label_lower and val < 0:
                        findings.append({"Sheet Name": sheetname, "Cell Reference": cell.coordinate, "Description": f"Label: {label}", "Category": "Sanity Check Failure", "Long Description": f"🔴 HIGH: Negative DSCR detected. Ratios should typically be positive."})
                    if "revenue" in label_lower and val < 0:
                        findings.append({"Sheet Name": sheetname, "Cell Reference": cell.coordinate, "Description": f"Label: {label}", "Category": "Sanity Check Failure", "Long Description": f"🔴 HIGH: Negative Revenue detected. Revenue should typically be positive."})

                orig_cell = ws[cell.coordinate]
                if orig_cell.data_type == 'f':
                    try:
                        r1c1 = a1_to_r1c1(orig_cell.value, orig_cell.row, orig_cell.column)
                        if len(str(orig_cell.value)) > 4000:
                            if val not in error_values:
                                findings.append({
                                    "Sheet Name": sheetname,
                                    "Cell Reference": cell.coordinate,
                                    "Description": "Deep Logic Check",
                                    "Category": "Auditability Risk",
                                    "Long Description": f"⚠️ MEDIUM: Mega-formula is technically functioning (no calculation error detected), but its complexity makes it an extreme auditability risk."
                                })
                    except: pass

        for row in ws.iter_rows():
            row_formulas = {}
            for cell in row:
                if cell.data_type == 'f':
                    try:
                        r1c1 = a1_to_r1c1(cell.value, cell.row, cell.column)
                        row_formulas[cell.coordinate] = r1c1
                    except: pass
            if len(row_formulas) > 5:
                counts = Counter(row_formulas.values())
                if len(counts) > 1:
                    dominant_r1c1, _ = counts.most_common(1)[0]
                    for coord, r1c1 in row_formulas.items():
                        if r1c1 != dominant_r1c1 and ws[coord].column >= 9:
                            findings.append({"Sheet Name": sheetname, "Cell Reference": coord, "Description": "Formula row", "Category": "Formula Pattern Break", "Long Description": f"⚠️ MEDIUM: Formula differs from dominant pattern in row. Expected R1C1: {dominant_r1c1} vs Actual R1C1: {r1c1}"})

            for cell in row:
                if cell.data_type == 'f':
                    formula = str(cell.value)
                    literals = re.findall(r'(?<![A-Z$])\b([0-9]+\.[0-9]+|[2-9][0-9]+|[0-9]{2,})\b', formula)
                    literals = [l for l in literals if l not in ['100', '12', '365', '52', '1', '0', '2', '4', '8000', '4000']]
                    if literals:
                        findings.append({"Sheet Name": sheetname, "Cell Reference": cell.coordinate, "Description": "Hard-coded literal", "Category": "Hard-Code in Formula", "Long Description": f"⚠️ MEDIUM: Formula contains hard-coded literal(s) {literals} which should likely be cell references."})

    return findings

if __name__ == "__main__":
    import sys
    import os
    if len(sys.argv) > 1:
        target_file = sys.argv[1]
    else:
        # Try to find any xlsx file in the parent directory if no argument is provided
        xlsx_files = [f for f in os.listdir('..') if f.endswith('.xlsx')]
        if xlsx_files:
            target_file = os.path.join('..', xlsx_files[0])
        else:
            print("No .xlsx file found to audit.")
            sys.exit(1)

    print(f"Auditing: {target_file}")
    results = full_audit(target_file)
    with open('full_audit_results.json', 'w') as f: json.dump(results, f)
