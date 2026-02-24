import openpyxl
import json
import re

def audit_hyperlinks(filename):
    wb = openpyxl.load_workbook(filename, data_only=True)
    findings = []

    # Load with formulas
    wb_form = openpyxl.load_workbook(filename, data_only=False)

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        ws_form = wb_form[sheetname]

        # Check formulas for HYPERLINK
        for row in ws_form.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and 'HYPERLINK' in str(cell.value).upper():
                    formula = str(cell.value)
                    # Simple regex to extract link from HYPERLINK("link", "friendly")
                    match = re.search(r'HYPERLINK\("([^"]+)"', formula, re.IGNORECASE)
                    if match:
                        link_target = match.group(1)
                        if link_target.startswith('#'):
                            # Internal link in formula
                            location = link_target[1:]
                            ref = cell.coordinate
                            if '!' in location:
                                t_sn, t_cr = location.split('!')
                                t_sn = t_sn.strip("'")
                                if t_sn not in wb.sheetnames:
                                    findings.append({
                                        "Sheet Name": sheetname, "Cell Reference": ref,
                                        "Description": f"HYPERLINK formula to {link_target}",
                                        "Category": "Broken Link - Missing Sheet",
                                        "Long Description": f"🔴 HIGH: HYPERLINK formula references sheet '{t_sn}' which does not exist."
                                    })
                        else:
                            # External link in formula
                            ref = cell.coordinate
                            if "example.com" in link_target:
                                findings.append({
                                    "Sheet Name": sheetname, "Cell Reference": ref,
                                    "Description": f"HYPERLINK formula to {link_target}",
                                    "Category": "Placeholder URL",
                                    "Long Description": f"🟡 LOW: External URL in formula contains obvious placeholder text."
                                })

        # openpyxl stores hyperlinks in ws._hyperlinks
        for rel in ws._hyperlinks:
            target = rel.target
            location = rel.location # e.g. 'Sheet1!A1'
            ref = rel.ref # e.g. 'A1'

            # Internal link
            if location:
                # Check if it's a sheet reference
                if '!' in location:
                    target_sheetname, target_cell_ref = location.split('!')
                    target_sheetname = target_sheetname.strip("'")
                    if target_sheetname not in wb.sheetnames:
                        findings.append({
                            "Sheet Name": sheetname,
                            "Cell Reference": ref,
                            "Description": f"Hyperlink to {location}",
                            "Category": "Broken Link - Missing Sheet",
                            "Long Description": f"🔴 HIGH: Hyperlink references sheet '{target_sheetname}' which does not exist."
                        })
                    else:
                        # Check if target cell is blank
                        target_ws = wb[target_sheetname]
                        try:
                            target_val = target_ws[target_cell_ref].value
                            if target_val is None or str(target_val).strip() == "":
                                findings.append({
                                    "Sheet Name": sheetname,
                                    "Cell Reference": ref,
                                    "Description": f"Hyperlink to {location}",
                                    "Category": "Link to Blank Cell",
                                    "Long Description": f"⚠️ MEDIUM: Hyperlink navigates to a valid cell, but the destination cell is blank."
                                })
                        except:
                            # Might be out of bounds or invalid ref
                            findings.append({
                                "Sheet Name": sheetname,
                                "Cell Reference": ref,
                                "Description": f"Hyperlink to {location}",
                                "Category": "Broken Link - Out of Bounds",
                                "Long Description": f"🔴 HIGH: Hyperlink targets a cell reference beyond the sheet's bounds or is invalid."
                            })
                else:
                    # Likely a named range or local cell
                    if location not in wb.defined_names:
                        # Try if it's a cell ref in the same sheet
                        try:
                            target_val = ws[location].value
                            if target_val is None or str(target_val).strip() == "":
                                findings.append({
                                    "Sheet Name": sheetname,
                                    "Cell Reference": ref,
                                    "Description": f"Hyperlink to {location}",
                                    "Category": "Link to Blank Cell",
                                    "Long Description": f"⚠️ MEDIUM: Hyperlink navigates to a valid cell, but the destination cell is blank."
                                })
                        except:
                             findings.append({
                                "Sheet Name": sheetname,
                                "Cell Reference": ref,
                                "Description": f"Hyperlink to {location}",
                                "Category": "Broken Link - Missing Name",
                                "Long Description": f"🔴 HIGH: Hyperlink targets a named range or cell '{location}' which does not exist."
                            })

            # External link
            if target and not location:
                if "example.com" in target or "xxx" in target:
                    findings.append({
                        "Sheet Name": sheetname,
                        "Cell Reference": ref,
                        "Description": f"Hyperlink to {target}",
                        "Category": "Placeholder URL",
                        "Long Description": f"🟡 LOW: External URL contains obvious placeholder text."
                    })
                if target.startswith("C:\\Users\\"):
                    findings.append({
                        "Sheet Name": sheetname,
                        "Cell Reference": ref,
                        "Description": f"Hyperlink to {target}",
                        "Category": "Non-Portable File Path",
                        "Long Description": f"⚠️ MEDIUM: File path is user-specific and will break for other users."
                    })

    return findings

if __name__ == "__main__":
    import sys
    import os
    target_file = sys.argv[1] if len(sys.argv) > 1 else None
    if not target_file:
        xlsx_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
        if xlsx_files: target_file = xlsx_files[0]

    if target_file:
        print(f"Auditing Hyperlinks: {target_file}")
        h_findings = audit_hyperlinks(target_file)

        # Merge with existing findings
        try:
            with open('full_audit_results.json', 'r') as f:
                all_findings = json.load(f)
        except:
            all_findings = []

        all_findings.extend(h_findings)

        with open('full_audit_results.json', 'w') as f:
            json.dump(all_findings, f)
        print(f"Added {len(h_findings)} hyperlink findings.")
