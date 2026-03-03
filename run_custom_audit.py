import os
import re
import string
from collections import defaultdict, Counter
import openpyxl
from openpyxl.utils import get_column_letter, coordinate_to_tuple
from spellchecker import SpellChecker

WORKBOOK_PATH = "CFG_Prep&Oship_Self-Auditing Model_Start.xlsx"
OUTPUT_DIR = "Output_v2"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Audit Report.md")

def create_output_dir():
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

# Utility for R1C1 conversion (basic approximation for pattern matching)
def to_r1c1(formula, row, col):
    if not formula.startswith('='):
        return formula

    def repl(match):
        # Match A1, $A$1, A$1, $A1
        col_str = match.group(1)
        row_str = match.group(2)

        c_idx = 0
        for char in col_str.replace('$', ''):
            c_idx = c_idx * 26 + (ord(char.upper()) - ord('A') + 1)
        r_idx = int(row_str.replace('$', ''))

        r_part = f"R{r_idx}" if '$' in row_str else f"R[{r_idx - row}]"
        c_part = f"C{c_idx}" if '$' in col_str else f"C[{c_idx - col}]"

        return r_part + c_part

    # Replace cell references
    r1c1 = re.sub(r'(\$?[A-Z]+)(\$?[0-9]+)', repl, formula)
    return r1c1

def group_cells(cells):
    # Basic grouping: just return comma separated if too complex, or range if contiguous
    if not cells: return ""
    if len(cells) == 1: return cells[0]
    # Simple grouping: just sort and join, a real contiguous range builder is complex
    return ", ".join(sorted(cells))

def audit_workbook():
    print("Loading workbook (this may take a minute)...")
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=False)
    try:
        wb_data = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    except:
        wb_data = wb # Fallback

    findings = []

    spell = SpellChecker()
    term_counts = Counter()

    # 1. Hyperlinks inventory
    hyperlinks = []

    # Pre-scan for Lingo (dominant terms)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            for val in row:
                if isinstance(val, str):
                    words = re.findall(r'\b[A-Za-z]+\b', val)
                    for w in words:
                        if len(w) > 3:
                            term_counts[w] += 1

    for sheet_name in wb.sheetnames:
        print(f"Auditing sheet: {sheet_name}")
        ws = wb[sheet_name]
        ws_data = wb_data[sheet_name]

        # Track formula patterns per row for Logic check
        row_formulas = defaultdict(list)

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                val = str(cell.value)
                val_data = str(ws_data[cell.coordinate].value) if ws_data[cell.coordinate].value is not None else ""
                c_row, c_col = cell.row, cell.column

                # SENTRY: Error Values
                if val_data in ["#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#SPILL!", "#CALC!", "#GETTING_DATA"]:
                    findings.append({
                        "sheet": sheet_name,
                        "cell": cell.coordinate,
                        "desc": "Error Value",
                        "category": "Calculation Error",
                        "long_desc": f"🔴 HIGH: Cell evaluates to error {val_data}"
                    })

                # HYPERLINKS
                if cell.hyperlink:
                    target = cell.hyperlink.target or cell.hyperlink.location
                    if target:
                        if "#" in target or cell.hyperlink.location:
                            # Internal link
                            loc = cell.hyperlink.location if cell.hyperlink.location else target.split("#")[-1]
                            if "!" in loc:
                                tgt_sheet = loc.split("!")[0].replace("'", "")
                                if tgt_sheet not in wb.sheetnames:
                                    findings.append({
                                        "sheet": sheet_name,
                                        "cell": cell.coordinate,
                                        "desc": "Broken Hyperlink",
                                        "category": "Broken Link – Missing Sheet",
                                        "long_desc": f"🔴 HIGH: Hyperlink targets a sheet '{tgt_sheet}' that does not exist."
                                    })
                if isinstance(cell.value, str) and 'HYPERLINK(' in cell.value.upper():
                    findings.append({
                        "sheet": sheet_name,
                        "cell": cell.coordinate,
                        "desc": "HYPERLINK Formula",
                        "category": "Hyperlink Check",
                        "long_desc": f"🟡 LOW: HYPERLINK formula found, ensure target exists: {val[:50]}..."
                    })

                # LOGIC & EFFICIENCY
                if val.startswith("="):
                    # Efficiency: Mega-formula
                    if len(val) > 500:
                        findings.append({
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "desc": "Formula > 500 chars",
                            "category": "Mega-Formula",
                            "long_desc": f"🔴 HIGH: Formula exceeds 500 characters ({len(val)} chars). This is bad practice."
                        })

                    # Logic: Hard-codes
                    hardcodes = re.findall(r'(?<![A-Z0-9$])\b\d+(?:\.\d+)?\b(?![A-Z0-9$])', val)
                    allowed = ["0", "1", "-1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "12", "52", "100", "365"]
                    suspicious = [h for h in hardcodes if h not in allowed]
                    if suspicious:
                        findings.append({
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "desc": "Hard-coded values",
                            "category": "Hard-Code in Formula",
                            "long_desc": f"⚠️ MEDIUM: Formulas contain hard-coded literal(s) {suspicious} instead of referencing input cells."
                        })

                    # Logic: Pattern break (Store for row-level check)
                    r1c1 = to_r1c1(val, c_row, c_col)
                    row_formulas[c_row].append((cell.coordinate, r1c1))

                # LINGO
                if isinstance(cell.value, str) and not val.startswith("="):
                    words = re.findall(r'\b[A-Za-z]+\b', val)
                    misspelled = spell.unknown(words)
                    # Filter out likely acronyms or formulas
                    misspelled = [w for w in misspelled if w.lower() == w and len(w) > 3]
                    if misspelled:
                        findings.append({
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "desc": "Potential Typo",
                            "category": "Typo",
                            "long_desc": f"🟡 LOW: Potential misspelling found: {misspelled}"
                        })

        # Process Logic Pattern Breaks per row
        for row_idx, cells_formulas in row_formulas.items():
            if len(cells_formulas) > 2:
                # Find dominant pattern
                patterns = [f for c, f in cells_formulas]
                counter = Counter(patterns)
                dominant, count = counter.most_common(1)[0]
                if count > 1 and count < len(patterns):
                    # We have a break
                    for c, f in cells_formulas:
                        if f != dominant:
                            findings.append({
                                "sheet": sheet_name,
                                "cell": c,
                                "desc": "Pattern Break",
                                "category": "Formula Pattern Break",
                                "long_desc": f"⚠️ MEDIUM: Cell formula breaks the dominant pattern in row {row_idx}."
                            })

        # SENTRY: Data Validations
        for dv in ws.data_validations.dataValidation:
            if dv.formula1 and "#REF!" in dv.formula1:
                findings.append({
                    "sheet": sheet_name,
                    "cell": dv.sqref.__str__(),
                    "desc": "Data Validation",
                    "category": "Invalid Validation",
                    "long_desc": f"🔴 HIGH: Data validation rule contains a broken reference (#REF!)."
                })

    # SENTRY: Defined Names
    try:
        if hasattr(wb, 'defined_names'):
            if isinstance(wb.defined_names, dict):
                items = wb.defined_names.items()
            elif hasattr(wb.defined_names, 'definedName'):
                items = [(dn.name, dn) for dn in wb.defined_names.definedName]
            else:
                items = []

            for name, dn in items:
                if hasattr(dn, 'value') and dn.value and "#REF!" in str(dn.value):
                    findings.append({
                        "sheet": "Workbook",
                        "cell": "Name Manager",
                        "desc": f"Name: {name}",
                        "category": "Dead Name",
                        "long_desc": f"🔴 HIGH: Named range '{name}' refers to a broken reference (#REF!)."
                    })
    except Exception as e:
        print(f"Error reading defined names: {e}")

    # MANAGER: Consolidate Report
    print("Writing Report...")
    create_output_dir()

    report = [
        "# Audit Report",
        "",
        "> Generated by Jules - Custom Python Audit Engine based on .Jules personas.",
        "",
        "| Sheet Name | Cell Reference | Description of the Location | Short Error Category | Long Description of Error |",
        "|---|---|---|---|---|"
    ]

    for f in findings:
        report.append(f"| {f['sheet']} | {f['cell']} | {f['desc']} | {f['category']} | {f['long_desc']} |")

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(report))

    print(f"Done! Report saved to {OUTPUT_FILE}")

if __name__ == "__main__":
    audit_workbook()
